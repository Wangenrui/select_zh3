# coding=utf-8
from openpyxl import *
import json
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "select_zh1.settings")
from django.db.models import Q
import sys
import django
django.setup()
from app.models import *
from django.forms.models import model_to_dict
from django.db import connection
from password import prpcrypt
import datetime
from decimal import Decimal
import uuid

pc = prpcrypt('boomboomboomboom')

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

def uuidToString(str):
    return str.replace('-', '')

def getProvinceid(company_id):
    try:
        company = Company.objects.get(id=company_id)
        if not company:
            return 0, 'Company不存在'
        else:
            return 1, company.company_province.id
    except:
        return 0, '数据库错误'

#周文超，教办管理员  新增

def supervisor_getSelectteacher(limit, offset,search,sort,order):

    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year

        cursor = connection.cursor()

        sql = """SELECT a.id,name,user_id,sex,tel,institution_name,major FROM app_select_sysuser a
              LEFT JOIN app_sex b ON a.sex_id=b.id
              LEFT JOIN app_institution c ON a.institution_id=c.id
              LEFT JOIN app_major d ON a.major_id=d.id
              WHERE (role_id=4 OR role_id=3) AND a.achieve_year= """+str(achieve_year)

        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR a.user_id LIKE \'%' + search + \
                  '%\' OR institution_name LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = """SELECT COUNT(*) FROM app_select_sysuser a
              LEFT JOIN app_sex b ON a.sex_id=b.id
              LEFT JOIN app_institution c ON a.institution_id=c.id
              WHERE (role_id=4 OR role_id=3) AND a.achieve_year=""" + str(achieve_year)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR a.user_id LIKE \'%' + search + \
                  '%\' OR institution_name LIKE \'%' + search + '%\')'
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'

def supervisor_getSelectstudent( limit, offset,search,sort,order):

    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        cursor = connection.cursor()
        sql = """SELECT a.id,a.name ,sex,a.user_id,a.tel,major,a.stu_class,a.score,f.name AS teachername,e.subject FROM app_select_sysuser a 
           LEFT JOIN app_sex b ON a.sex_id=b.id 
           LEFT JOIN app_major c on a.major_id=c.id 
           LEFT JOIN app_application_state d ON a.user_id=d.selected_stu_no 
           LEFT JOIN app_topic e ON d.topic= e.id 
           LEFT JOIN app_select_sysuser f ON e.tea_no=f.user_id 
           where a.role_id=1 AND a.achieve_year=""" + str(achieve_year)

        if search != '':
           sql = sql + ' AND (a.name LIKE \'%' + search + \
                 '%\' OR a.user_id LIKE \'%' + search + \
                 '%\' OR a.stu_class LIKE \'%' + search + \
                 '%\' OR major LIKE \'%' + search +'%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = """SELECT COUNT(*) FROM app_select_sysuser a 
           LEFT JOIN app_sex b ON a.sex_id=b.id 
           LEFT JOIN app_major c on a.major_id=c.id 
           LEFT JOIN app_application_state d ON a.user_id=d.selected_stu_no 
           LEFT JOIN app_topic e ON d.topic= e.id 
           LEFT JOIN app_select_sysuser f ON e.tea_no=f.user_id 
           where a.role_id=1 AND a.achieve_year=""" + str(achieve_year)
        if search != '':
           sql = sql + ' AND (a.name LIKE \'%' + search + \
                 '%\' OR a.user_id LIKE \'%' + search + \
                 '%\' OR a.stu_class LIKE \'%' + search + \
                 '%\' OR major LIKE \'%' + search +'%\')'
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'


def supervisor_getUserInfo(userid):
    try:
        dict = {}
        id = select_sysUser.objects.get(user_id=userid)
        uid = str(id.id).replace('-', '')
        cursor = connection.cursor()
        cursor.execute("""
                  SELECT a.id,name,user_id,sex,tel,institution_name,d.role_name FROM app_select_sysuser a
                     LEFT JOIN app_sex b ON a.sex_id=b.id
                     LEFT JOIN app_institution c ON a.institution_id=c.id
                     LEFT join app_select_role d ON a.role_id = d.id                     
                     WHERE a.user_id = %s
                        """, [userid])
        dict['data'] = dictfetchall(cursor)

        return 1, dict
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def supervisor_addAchieveyear(achieve_year, nextyear):
    try:

        suer = Date_setting.objects.filter(year=achieve_year)

        if len(suer) > 0:

            subject = Date_setting.objects.get(year=achieve_year)

            if subject.activation == 1:
                return 0, achieve_year + '年份已开启'

            else:

                activateyears = Date_setting.objects.filter(activation=1)
                nextyears = Date_setting.objects.filter(year=nextyear)

                for activateyear in activateyears:
                    activateyear.activation = 0
                    activateyear.save()

                subject.activation = 1
                subject.save()
                if len(nextyears) <= 0:
                    obj = Date_setting(
                        year=nextyear,
                        activation=0
                    )
                    obj.save()

                return 0, achieve_year + '年份开启成功'

        else:

            activationyears = Date_setting.objects.filter(activation=1)
            if len(activationyears) <= 0:
                obj = Date_setting(
                    year=achieve_year,
                    activation=1
                )
                obj.save()
                obj1 = Date_setting(
                    year=nextyear,
                    activation=0
                )
                obj1.save()


            else:
                for activationyear in activationyears:
                    activationyear.activation = 0
                    activationyear.save()
                    obj = Date_setting(
                        year=achieve_year,
                        activation=1
                    )
                    obj.save()
                    obj1 = Date_setting(
                        year=nextyear,
                        activation=0
                    )
                    obj1.save()

            return 1, achieve_year + '年份开启成功'
    except:
        return 0
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def supervisor_getYear():
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_date_setting """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def supervisor_getDatebyid(eid):
    id = eid.replace('-', '')

    try:
        cursor = connection.cursor()
        sql = '''SELECT *
                 FROM app_date_setting               
                 WHERE id=\''''

        sql = sql + str(id) + '\''
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def supervisor_setDate(
        year,
        av_start,
        av_end,
        write_sub_start,
        write_sub_end,
        stu_select_start,
        stu_select_end,
        tea_select_start
):
    try:
        obj = Date_setting.objects.get(id=year)
        obj.av_start = av_start
        obj.av_end = av_end
        obj.write_sub_s = write_sub_start
        obj.write_sub_e = write_sub_end
        obj.select1_start = stu_select_start
        obj.select1_end = stu_select_end
        obj.select3_start = tea_select_start

        obj.save()
        return 1, '时间设置成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def supervisor_getInstitution():
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_institution """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def supervisor_addstudentMessage(
        stu_name,
        sex_id,
        stu_id,
        stu_major_id,
        stu_class,
        stu_score,
        tel

):
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        year = achieveyear.year
        count = select_sysUser.objects.filter(user_id=stu_id)
        psw = pc.encrypt('111111')
        if len(count) > 0:
            return 0, '学生已存在'
        else:
            obj = select_sysUser(
                name=stu_name,
                sex_id=sex_id,
                user_id=stu_id,
                major_id=stu_major_id,
                stu_class=stu_class,
                score=stu_score,
                tel=tel,
                password=psw,
                achieve_year=year,
                role_id=1
            )
            obj.save()

            return 1, '添加成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info



def supervisor_updateStudentmessage(
        id,
        stu_name,
        sex_id,
        stu_id,
        stu_major_id,
        stu_class,
        stu_score,
        tel

):
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        year = achieveyear.year
        count = select_sysUser.objects.filter(id=id)
        if len(count) == 0:
            return 0, '学生不存在'
        else:
            object = select_sysUser.objects.get(id=id)

            object.name = stu_name
            object.user_id = stu_id
            object.sex_id = sex_id
            object.major_id = stu_major_id
            object.score = stu_score
            object.stu_class = stu_class
            object.tel = tel
            object.achieve_year = year

            object.save()
            return 1, '更新成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def supervisor_getTeacherbyid(eid):
    id = eid.replace('-', '')
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        cursor = connection.cursor()
        sql = '''SELECT * FROM app_select_sysuser WHERE  id='''+str(id)+''' AND achieve_year = '''+str(achieve_year)
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def supervisor_delTeacherbyid(id):
    try:
        select_sysUser.objects.get(id=id).delete()
        return 1, '删除成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def supervisor_addteacherMessage(
        teacher_name ,
        teacher_id,
        sex_id,
        telephone,
        institution_id,
        major_id
):
    try:
        achieve_year = Date_setting.objects.get(activation=1)
        count = select_sysUser.objects.filter(user_id=teacher_id)
        if len(count) > 0:
            return 0, '该教师已存在'
        else:
            psw = pc.encrypt('111111')
            obj = select_sysUser(
                name=teacher_name,
                user_id=teacher_id,
                sex_id=sex_id,
                tel=telephone,
                password=psw,
                achieve_year=achieve_year.year,
                role_id=4,
                institution_id=institution_id,
                major_id=major_id
            )
            obj.save()

            return 1, '添加成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info



def supervisor_updateTeachermessage(
        id,
        teacher_name,
        teacher_id,
        sex_id,
        telephone,
        institution_id,
        major_id
):
    try:
        count = select_sysUser.objects.filter(id=id)
        if len(count) == 0:
            return 0, '教师不存在'
        else:
            object = select_sysUser.objects.get(id=id)

            object.name = teacher_name
            object.user_id = teacher_id
            object.sex_id = sex_id
            object.tel = telephone
            object.institution_id = institution_id
            object.major_id = major_id

            object.save()
            return 1, '更新成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info




def supervisor_getStudentbyid(eid):
    id = eid.replace('-', '')
    try:
        cursor = connection.cursor()
        sql = '''SELECT * 
                 FROM app_select_sysuser               
                 WHERE role_id=1 AND id=\''''

        sql = sql + str(id) + '\''
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def supervisor_delStudentbyid(id):
    try:
        select_sysUser.objects.get(id=id).delete()
        return 1, '删除成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info







def supervisor_saveSelectFileStudent(filename):
    try:
        print filename
        Name = str(filename).split('.')
        if Name[-1]!='xlsx':
            return 0,'请上传后缀为xlsx的文件！'
        file = File1()
        file.name = filename
        file.save()
        wb = load_workbook(filename)

        sheet_names = wb.get_sheet_names()  #
        ws = wb.get_sheet_by_name(sheet_names[0])  # 读第一张表为：index[0]
        max_row = ws.max_row
        psw = pc.encrypt('111111')
        Class.objects.all().delete()
        class_name = []

        #读表中的专业，更新major表
        for i in range(2, max_row+1):
            H = []
            for j in ws[i]:
                H.append(j.value)

            temp1 = Major.objects.filter(major=H[2])

            if temp1:
                print '1'
            else:
                obj2 = Major(
                    major = H[2]
                )
                obj2.save()


        for i in range(2, max_row+1):
            L = []
            for j in ws[i]:
                L.append(j.value)
            class_name.append(L[3])
            user = select_sysUser.objects.filter(user_id__exact=L[0])
            if user:
                test = select_sysUser.objects.get(user_id__exact=L[0])
                test.name = L[1]
                test.password=psw
                test.major_id = L[2]
                test.stu_class = L[3]
                test.tel = L[4]
                test.achieve_year = L[5]
                test.role_id = 1

                test.sex_id = Sex.objects.get(sex__exact=L[6])
                test.score=L[7]
                test.save()
            else:
                User_id = L[0]
                Name = L[1]
                major = Major.objects.get(major = L[2])
                major_id=major.id
                Stu_class = L[3]
                Tel = L[4]
                Achieve_year = L[5]

                Sex_id = Sex.objects.get(sex__exact=L[6])
                score = L[7]
                obj = select_sysUser(
                            user_id=User_id,
                            name=Name,
                            major_id=major_id,
                            password=psw,
                            stu_class=Stu_class,
                            tel=Tel,
                            achieve_year=Achieve_year,

                            role_id=1,
                            sex_id=Sex_id.id,
                            score = score
                )

                obj.save()

        #读表中的班级，更新class表
        newClass_name = []
        for x in class_name:
            if x not in newClass_name:
                newClass_name.append(x)
        newClass_name.sort()

        for y in newClass_name:
            obj1 = Class(
                class_name = y,
            )

            obj1.save()


        return 1, '文件上传成功'

    except:

        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def supervisor_saveSelectFileTeacher(filename):
    try:
        file = File1()
        file.name = filename
        file.save()
        print filename
        wb = load_workbook(filename)
        sheet_names = wb.get_sheet_names()  #
        ws = wb.get_sheet_by_name(sheet_names[0])  # 读第一张表为：index[0]
        max_row = ws.max_row  # 11     <-最大行数
        psw = pc.encrypt('111111')
        institution = []

        # 读表中的专业，更新major表
        for i in range(2, max_row + 1):
            H = []
            for j in ws[i]:
                H.append(j.value)

            temp1 = Major.objects.filter(major=H[7])

            if temp1:
                print '1'
            else:
                obj2 = Major(
                    major=H[7]
                )
                obj2.save()

        # 读表中的研究所，更新institution表
        for i in range(2, max_row + 1):

            H = []
            for j in ws[i]:

                H.append(j.value)
            temp = Institution.objects.filter(institution_name__exact=H[3])
            if temp:
                print ''
            else:
                obj2 = Institution(
                    institution_name=H[3]
                )
                obj2.save()

        for i in range(2, max_row+1):

            L = []
            for j in ws[i]:
                L.append(j.value)
            user = select_sysUser.objects.filter(user_id__exact=L[0])
            if user:
                test = select_sysUser.objects.get(user_id__exact=L[0])
                test.name = L[1]
                test.password=psw
                test.stu_class = '无'
                test.tel = L[2]
                test.achieve_year = L[5]
                if L[6]=='是':
                    test.role_id = 3
                else:
                    test.role_id = 4
                temp = Major.objects.get(major=L[7])
                test.major_id=temp.id
                test.institution_id = Institution.objects.get(institution_name__exact=L[3])
                test.sex_id = Sex.objects.get(sex__exact=L[4])
                test.save()

            else:
                User_id = L[0]
                Name = L[1]
                Tel = L[2]
                Institution_id = Institution.objects.get(institution_name__exact=L[3])
                if L[6]=='是':
                    role_id = 3
                else:
                    role_id = 4
                temp = Major.objects.get(major=L[7])

                Sex_id = Sex.objects.get(sex__exact=L[4])
                obj = select_sysUser(
                    user_id=User_id,
                    name=Name,
                    password=psw,
                    stu_class='无',
                    tel=Tel,
                    achieve_year=L[5],
                    institution_id=Institution_id.id,
                    role_id=role_id,
                    sex_id=Sex_id.id,
                    major_id=temp.id
                )
                obj.save()

        return 1, '文件上传成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info



def supervisor_initStudent():
    try:
        students = Application.objects.all()
        #selectedstudents = Application_state.objects.all()
        for student in students:
            student.submit = 0
            student.save()

        return 1, '未成功选题的同学已初始化'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


##周文超





##张煜，学生    新增


def student_getApplication(stu_no):
    try:
        cursor = connection.cursor()
        sql = "SELECT * FROM app_application WHERE app_application.stu_no="+str(stu_no)
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'




def student_getTopic(stuMajor):

    try:
        cursor = connection.cursor()
        sql = '''SELECT a.id,a.subject,b.institution_id,b.name FROM app_topic a
                 LEFT JOIN app_select_sysuser b ON a.tea_no=b.user_id
                 LEFT JOIN app_date_setting c ON a.year=c.year
                 WHERE a.id not in (SELECT app_application_state.topic FROM app_application_state)
                 AND c.activation=1 
                 AND a.major_id=''' +stuMajor

        cursor.execute(sql)
        return 1,dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def student_getTopicIntroduction(topic_id):
    try:
        topic=Topic.objects.get(id=topic_id)
        return 1, topic.introduction
    except:
        return 0, ''



def student_getFile(limit, offset,subject):
    try:
        cursor = connection.cursor()
        sql = '''SELECT a.id,a.file_position,b.subject FROM app_file a
                 LEFT JOIN app_topic b ON b.id=a.subject
                 WHERE a.subject = '''+"'"+str(subject)+"'"
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        sql='''SELECT COUNT(*) FROM app_file WHERE subject ='''+"'"+str(subject)+"'"
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'] , 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'



def student_addApplication(
    stu_no,
    id1,
    volunteer_no1,
    volunteer_topic_no_id1,
    id2,
    volunteer_no2,
    volunteer_topic_no_id2,
    id3,
    volunteer_no3,
    volunteer_topic_no_id3,
    id4,
    volunteer_no4,
    volunteer_topic_no_id4,
    id5,
    volunteer_no5,
    volunteer_topic_no_id5,
    id6,
    volunteer_no6,
    volunteer_topic_no_id6
):

    try:

        obj = Application(
            id=id1,
            stu_no=stu_no,
            volunteer_no=volunteer_no1,
            volunteer_topic_no_id=volunteer_topic_no_id1)
        obj.save()
        obj = Application(
            id=id2,
            stu_no=stu_no,
            volunteer_no=volunteer_no2,
            volunteer_topic_no_id=volunteer_topic_no_id2)
        obj.save()
        obj = Application(
            id=id3,
            stu_no=stu_no,
            volunteer_no=volunteer_no3,
            volunteer_topic_no_id=volunteer_topic_no_id3)
        obj.save()
        obj = Application(
            id=id4,
            stu_no=stu_no,
            volunteer_no=volunteer_no4,
            volunteer_topic_no_id=volunteer_topic_no_id4)
        obj.save()
        obj = Application(
            id=id5,
            stu_no=stu_no,
            volunteer_no=volunteer_no5,
            volunteer_topic_no_id=volunteer_topic_no_id5)
        obj.save()
        obj = Application(
            id=id6,
            stu_no=stu_no,
            volunteer_no=volunteer_no6,
            volunteer_topic_no_id=volunteer_topic_no_id6)
        obj.save()
        return 1, '保存成功'

    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info







def student_getState(stu_no):
    try:
        dict = {}
        subject=Application_state.objects.get(selected_stu_no=stu_no)
        topic_id=subject.topic
        tea=Topic.objects.get(id=topic_id)
        tea=tea.tea_no
        cursor = connection.cursor()
        cursor.execute("""
                 SELECT  a.topic,b.subject
                 FROM app_application_state a  
                 LEFT JOIN app_topic b ON a.topic=b.id
                 WHERE a.selected_stu_no=%s
                         """, [stu_no])
        dict['data'] = dictfetchall(cursor)

        cursor.execute("""
                 SELECT  a.introduction
                 FROM app_topic a  
                 WHERE a.id=%s
                         """, [topic_id])
        dict['info'] = dictfetchall(cursor)

        cursor.execute("""
                 SELECT  a.name,a.tel
                 FROM app_select_sysuser a  
                 WHERE a.user_id=%s
                         """, [tea])
        dict['teacher'] = dictfetchall(cursor)

        cursor.execute("""
                    SELECT  a.select1_start,a.select1_end,a.select2_start,a.select2_end,a.select3_start,a.select3_end FROM app_date_setting a
                    LEFT JOIN app_select_sysuser b ON  a.year=b.achieve_year 
                    WHERE a.activation=1 AND b.user_id=%s
                            """, [stu_no])
        dict['date']=dictfetchall(cursor)

        return 1, dict
    except:
        return 0, '数据库错误'

def student_getDate(stu_no):
    try:
        dict = {}
        cursor = connection.cursor()
        cursor.execute("""
                        SELECT  a.year,a.select1_start,a.select1_end,a.select2_start,a.select2_end,a.select3_start,a.select3_end FROM app_date_setting a
                        LEFT JOIN app_select_sysuser b ON  a.year=b.achieve_year 
                        WHERE a.activation=1 AND b.user_id=%s
                                """, [stu_no])
        dict['date'] = dictfetchall(cursor)


        return 1, dict
    except:
        return 0, '数据库错误'

####张煜







##韩佳琦 ， 教学所长   新增

def director_getTime(tea_no):
    try:
        dict = {}

        cursor = connection.cursor()

        cursor.execute("""
                            SELECT  * FROM app_date_setting 
                                    WHERE activation = 1 
                                     """)
        dict['date']=dictfetchall(cursor)

        return 1, dict
    except:
        return 0, '数据库错误'


def director_getjianjie(institution_id,topic_id):
    try:

        cursor = connection.cursor()
        sql = "SELECT  * FROM app_topic WHERE id = '" + topic_id + "'"

        cursor.execute(sql)

        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def director_getnum(tea_id):
    aaa = select_sysUser.objects.get(id=tea_id)
    teano = aaa.user_id
    thisyear = Date_setting.objects.get(activation=1)
    achieveyear = thisyear.year
    try:

        cursor = connection.cursor()
        sql = "SELECT  * FROM app_teacher_topic_num  WHERE tea_no = '" + teano + "'"
        sql = sql +' and year=' + str(achieveyear)
        cursor.execute(sql)

        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def director_getnum1(tea_id):
    try:

        cursor = connection.cursor()
        sql = "SELECT  * FROM app_teacher_topic_num  WHERE tea_no = '" + tea_id + "'"

        cursor.execute(sql)

        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'
def director_getInsti():
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_institution """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def test():
    return True, 1




def director_getTeacher(institution_id):
    try:
        cursor = connection.cursor()
        sql = "SELECT  * FROM app_date_setting WHERE activation = 1"
        cursor.execute(sql)
        year = dictfetchall(cursor)[0]
        cursor = connection.cursor()
        sql = "SELECT * FROM app_select_sysuser a  WHERE (a.role_id=4 or role_id=3 )" \
              "and achieve_year = '" + year['year'] + "'"
        sql = sql + ' AND a.institution_id=\'' + institution_id + '\''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'
def director_getTopic(institution_id,province_id):

    try:
        cursor = connection.cursor()
        sql = "SELECT  * FROM app_select_sysuser WHERE id = '" + province_id + "'"
        cursor.execute(sql)
        teacher = dictfetchall(cursor)[0]
        cursor = connection.cursor()
        sql = "SELECT  * FROM app_date_setting WHERE activation = 1"
        cursor.execute(sql)
        year = dictfetchall(cursor)[0]
        cursor = connection.cursor()
        sql = "SELECT id,subject FROM app_topic " \
              "WHERE id not in (SELECT app_application_state.topic FROM app_application_state ) "\
            "  and year = '" + year['year'] + "'"\
              "  and tea_no = '" + teacher['user_id'] + "'"

        cursor.execute(sql)

        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'




def director_getClass(institution_id):
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_class """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def director_getTeacc(institution_id):
    try:
        cursor = connection.cursor()
        sql = "SELECT  * FROM app_date_setting WHERE activation = 1"
        cursor.execute(sql)
        year = dictfetchall(cursor)[0]
        cursor = connection.cursor()
        sql = "SELECT * FROM app_select_sysuser a  WHERE (a.role_id=4 or role_id=3) " \
              "  and achieve_year = '" + year['year'] + "'"
        sql = sql + ' AND a.institution_id=\'' + institution_id + '\''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'



def director_getSelectstudent(institution_id,tea_no,class_id,stustu,limit,offset,search,sort,order):

    institution = institution_id.replace('-', '')
    try:
        '''
        "SELECT a.id,a.name ,f.name AS Bname ,f.user_id AS Bno,b.sex ,a.user_id,a.stu_class,d.topic,e.subject FROM app_select_sysuser a " \
              "LEFT JOIN app_sex b ON a.sex_id=b.id " \
              "LEFT JOIN app_institution c ON a.institution_id=c.id " \
              "LEFT JOIN app_application_state d ON a.user_id=d.selected_stu_no " \
              "LEFT JOIN app_topic e ON d.topic=e.id " \
              "LEFT JOIN app_select_sysuser AS f ON e.tea_no=f.user_id " \
              "LEFT JOIN app_class h ON a.stu_class = h.class_name where a.role_id= 1 and a.institution_id = '" + institution+"'" \
              "and a.achieve_year = '" + year['year'] + "'"
                      if tea_no != '' and tea_no != '-1':
         sql = sql + ' AND f.id=' + str(tea_no)
        if stustu != '' and stustu != '1':
         sql = sql + ' AND e.subject IS NULL'
        if class_id != '' and class_id != '-1':
         sql = sql + ' AND h.id=' + str(class_id)
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        
        
        a:app_select_sysuser  b:app_topic  c:app_select_sysuser  d:app_application_state
        SELECT a.id,a.name AS Bname,b.subject,c.name,c.user_id,e.sex,c.stu_class FROM app_select_sysuser a
        LEFT JOIN app_topic b ON b.tea_no=a.user_id
        LEFT JOIN app_application_state d ON d.topic=b.id
        LEFT JOIN app_select_sysuser c ON c.user_id = d.selected_stu_no
        LEFT JOIN app_sex e ON e.id=c.sex_id
        WHERE a.achieve_year='2017' and a.institution_id=1 and b.year='2017' 
        '''
        cursor = connection.cursor()
        sql = "SELECT  * FROM app_date_setting WHERE activation = 1"
        cursor.execute(sql)
        year = dictfetchall(cursor)[0]
        cursor = connection.cursor()

        sql = '''SELECT a.id,a.name AS Bname,b.subject,c.name,c.user_id,e.sex,c.stu_class FROM app_select_sysuser a
        LEFT JOIN app_topic b ON b.tea_no=a.user_id
        LEFT JOIN app_application_state d ON d.topic=b.id
        LEFT JOIN app_select_sysuser c ON c.user_id = d.selected_stu_no
        LEFT JOIN app_sex e ON e.id=c.sex_id
        LEFT JOIN app_class f ON c.stu_class = f.class_name
        WHERE  a.institution_id='''+institution+" and b.year="+"'"+year['year']+"'"+" and a.achieve_year="+"'"+year['year']+"'"+" and c.achieve_year="+"'"+year['year']+"'"
        if tea_no != '' and tea_no != '-1':
         sql = sql + ' AND a.id=' + str(tea_no)
        if class_id != '' and class_id != '-1':
         sql = sql + ' AND f.id=' + str(class_id)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR b.subject LIKE \'%' + search + \
                  '%\' OR c.name LIKE \'%' + search + \
                  '%\' OR e.sex LIKE \'%' + search + \
                  '%\' OR c.user_id LIKE \'%' + search + \
                  '%\' OR c.stu_class LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)

        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        print sql
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        sql = '''SELECT COUNT(*) FROM app_select_sysuser a
        LEFT JOIN app_topic b ON b.tea_no=a.user_id
        LEFT JOIN app_application_state d ON d.topic=b.id
        LEFT JOIN app_select_sysuser c ON c.user_id = d.selected_stu_no
        LEFT JOIN app_sex e ON e.id=c.sex_id
        LEFT JOIN app_class f ON c.stu_class = f.class_name
        WHERE  a.institution_id='''+institution+" and b.year="+"'"+year['year']+"'"+" and a.achieve_year="+"'"+year['year']+"'"+" and c.achieve_year="+"'"+year['year']+"'"
        if tea_no != '' and tea_no != '-1':
            sql = sql + ' AND a.id=' + str(tea_no)
        if class_id != '' and class_id != '-1':
            sql = sql + ' AND f.id=' + str(class_id)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR b.subject LIKE \'%' + search + \
                  '%\' OR c.name LIKE \'%' + search + \
                  '%\' OR e.sex LIKE \'%' + search + \
                  '%\' OR c.user_id LIKE \'%' + search + \
                  '%\' OR c.stu_class LIKE \'%' + search + '%\')'
        print sql
        cursor.execute(sql)
        num=dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows':select_out}
        return data
    except:
        return 0, '数据库错误'



def director_getStudentnobyid(eid):


    id = eid.replace('-', '')
    try:
        cursor = connection.cursor()

        sql = "SELECT  * FROM app_select_sysuser WHERE role_id =  1 AND id='" + id + "'"
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def director_updateTopic(
        id,
        stu_id,
        submit,
        topic_id):
    try:

        obj = Application_state(

            topic=topic_id,
            selected_stu_no=stu_id,
            submit=submit,
        )
        obj.save()
        return 1, '指派成功'
    except:
        import sys

        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def director_updateNum(
        teacc,
        topic_num):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    aaa = select_sysUser.objects.get(id=teacc)
    teano = aaa.user_id

    try:

        obj = Teacher_topic_num(
            year = achieve_year,
            tea_no=teano,
            topic_num=topic_num,
        )
        obj.save()
        return 1, '保存成功'
    except:
        import sys

        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info




def director_xiugaitopicNum(

        teacc,
        topic_num

):

    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year



    try:

        temp = Teacher_topic_num.objects.filter(year = achieve_year ,tea_no = teacc)
        if temp :
            temp1 = Teacher_topic_num.objects.get(year=achieve_year, tea_no=teacc)

            temp1.topic_num = topic_num


            temp1.save()
            return 1, '添加成功'
        else:
            obj = Teacher_topic_num(
                tea_no=teacc,
                topic_num=topic_num,
                year=achieve_year

            )
            obj.save()
            return 1, '添加成功'






    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info
def director_addtopicNum(

                teacc,
                topic_num

            ):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year

    try:


        obj = Teacher_topic_num(
                tea_no=teacc,
                topic_num=topic_num,
                year=achieve_year

        )
        obj.save()
        return 1, '添加成功'



    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def director_Addintro(
                userid,
                teacc,
                subject,
                introduction,
                subject_property_id,
                other_introduction,
                combine_actual,
                company_name,
                implementation_state,
                english_title,

            ):

    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    teamajor = select_sysUser.objects.get(user_id=teacc)
    major = teamajor.major_id

    try:
        print '上传题目'

        if combine_actual=='是':
            combine_actual='√'
        if subject_property_id =='研究论文':

            obj = Topic(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                research_papers='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major
            )
            obj.save()
            ID = obj.id

            obj1 = Topic1(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                research_papers='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major,
                topic = ID
            )
            obj1.save()

            return 1, '添加成功'

        if subject_property_id == '工程设计':

            obj = Topic(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                engineering_design='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major
            )
            obj.save()
            ID = obj.id

            obj1 = Topic1(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                engineering_design='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major,
                topic = ID
            )
            obj1.save()
            return 1, '添加成功'

        if subject_property_id == '项目报告':

            obj = Topic(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                project_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major

            )
            obj.save()
            ID = obj.id

            obj1 = Topic1(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                project_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major,
                topic = ID

            )
            obj1.save()
            return 1, '添加成功'

        if subject_property_id == '综述报告':

            obj = Topic(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                summary_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major
            )
            obj.save()
            ID = obj.id
            obj1 = Topic1(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                summary_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major,
                topic = ID
            )
            obj1.save()
            return 1, '添加成功'

        if subject_property_id == '其他':

            obj = Topic(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                other='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major
            )
            obj.save()
            ID = obj.id
            obj1 = Topic1(
                tea_no=teacc,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                other='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=major,
                topic = ID
            )
            obj1.save()
            return 1, '添加成功'

    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def director_getStuInfo(userid):
    try:
        dict = {}
        cursor = connection.cursor()
        cursor.execute("""
                      SELECT a.user_id,a.name,a.tel,b.sex  
                      FROM app_select_sysuser a 
                      LEFT JOIN app_sex b  ON b.id=a.sex_id 
                      WHERE a.user_id=%s
                            """, [userid])
        dict['data'] = dictfetchall(cursor)

        cursor.execute("""
                      SELECT b.institution_name 
                      FROM app_select_sysuser a 
                      LEFT JOIN app_institution b ON a.institution_id=b.id  
                      WHERE a.user_id = %s
                             """, [userid])
        dict['institution'] = dictfetchall(cursor)

        return 1, dict
    except:
        return 0, '数据库错误'

def getTeaToopic(limit, offset, tea_no,userid,search,sort,order):

    try:
        user=select_sysUser.objects.get(user_id=userid)
        institutionid=user.institution_id
        cursor = connection.cursor()
        sql = '''SELECT  a.id,a.subject,a.introduction,a.year,c.name,c.stu_class,c.user_id ,d.sex,e.year,e.activation,f.major FROM app_topic a
                 LEFT JOIN app_select_sysuser c ON c.user_id=a.tea_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 WHERE e.activation=1 '''
        sql = sql + ' AND c.institution_id=' + str(institutionid)
        if tea_no != '' and tea_no != '-1':
         sql = sql + ' AND c.id=' + str(tea_no)
        if search != '':
            sql = sql + ' AND (c.name LIKE \'%' + search + \
                  '%\' OR a.subject LIKE \'%' + search + \
                  '%\' OR f.major LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)
        sql = sql  + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = '''SELECT COUNT(*)  FROM app_topic a
                 LEFT JOIN app_select_sysuser c ON c.user_id=a.tea_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 WHERE e.activation=1 '''
        sql = sql + ' AND c.institution_id=' + str(institutionid)
        if tea_no != '' and tea_no != '-1':
         sql = sql + ' AND c.id=' + str(tea_no)
        if search != '':
            sql = sql + ' AND (c.name LIKE \'%' + search + \
                  '%\' OR a.subject LIKE \'%' + search + \
                  '%\' OR f.major LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + '%\')'
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'

def xiugaiTopic(
        id,
        teacc,
        subject,
        introduction,
        subject_property_id,
        other_introduction,
        combine_actual,
        company_name,
        implementation_state,
        english_title,

):

    try:

        count = Topic.objects.filter(id=id)
        if len(count) == 0:
            return 0, '题目不存在'
        if combine_actual=='是':
            combine_actual='√'
        if subject_property_id=='研究论文':
           object = Topic.objects.get(id=id)
           object.tea_no = teacc
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = '√'
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.save()

           object1 = Topic1.objects.get(topic=id)
           object1.tea_no = teacc
           object1.subject = subject
           object1.introduction = introduction
           object1.subject_property = subject_property_id
           object1.research_papers = '√'
           object1.engineering_design = ''
           object1.project_report = ''
           object1.summary_report = ''
           object1.other = ''
           object1.other_introduction = other_introduction
           object1.combine_actual = combine_actual
           object1.company_name = company_name
           object1.implementation_state = implementation_state
           object1.english_title = english_title
           object1.save()


           return 1, '更新成功'

        if subject_property_id=='工程设计':
           object = Topic.objects.get(id=id)
           object.tea_no = teacc
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = '√'
           object.project_report = ''
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.save()

           object1 = Topic1.objects.get(topic=id)
           object1.tea_no = teacc
           object1.subject = subject
           object1.introduction = introduction
           object1.subject_property = subject_property_id
           object1.research_papers = ''
           object1.engineering_design = '√'
           object1.project_report = ''
           object1.summary_report = ''
           object1.other = ''
           object1.other_introduction = other_introduction
           object1.combine_actual = combine_actual
           object1.company_name = company_name
           object1.implementation_state = implementation_state
           object1.english_title = english_title
           object1.save()
           return 1, '更新成功'

        if subject_property_id=='项目报告':
           object = Topic.objects.get(id=id)
           object.tea_no = teacc
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = '√'
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.save()

           object1 = Topic1.objects.get(topic=id)
           object1.tea_no = teacc
           object1.subject = subject
           object1.introduction = introduction
           object1.subject_property = subject_property_id
           object1.research_papers = ''
           object1.engineering_design = ''
           object1.project_report = '√'
           object1.summary_report = ''
           object1.other = ''
           object1.other_introduction = other_introduction
           object1.combine_actual = combine_actual
           object1.company_name = company_name
           object1.implementation_state = implementation_state
           object1.english_title = english_title
           object1.save()
           return 1, '更新成功'

        if subject_property_id=='综述报告':
           object = Topic.objects.get(id=id)
           object.tea_no = teacc
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = '√'
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.save()

           object1 = Topic1.objects.get(topic=id)
           object1.tea_no = teacc
           object1.subject = subject
           object1.introduction = introduction
           object1.subject_property = subject_property_id
           object1.research_papers = ''
           object1.engineering_design = ''
           object1.project_report = ''
           object1.summary_report = '√'
           object1.other = ''
           object1.other_introduction = other_introduction
           object1.combine_actual = combine_actual
           object1.company_name = company_name
           object1.implementation_state = implementation_state
           object1.english_title = english_title
           object1.save()
           return 1, '更新成功'

        if subject_property_id=='其他':
           object = Topic.objects.get(id=id)
           object.tea_no = teacc
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = ''
           object.other = '√'
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.save()

           object1 = Topic1.objects.get(topic=id)
           object1.tea_no = teacc
           object1.subject = subject
           object1.introduction = introduction
           object1.subject_property = subject_property_id
           object1.research_papers = ''
           object1.engineering_design = ''
           object1.project_report = ''
           object1.summary_report = ''
           object1.other = '√'
           object1.other_introduction = other_introduction
           object1.combine_actual = combine_actual
           object1.company_name = company_name
           object1.implementation_state = implementation_state
           object1.english_title = english_title
           object1.save()
           return 1, '更新成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def director_delTopic(id):
    try:
        cursor = connection.cursor()
        sql = '''DELETE FROM app_topic WHERE id ='''+id
        cursor.execute(sql)
        sql = '''DELETE FROM app_topic1 WHERE topic =''' + id
        cursor.execute(sql)
        return 1, '删除成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info




def director_uptopic(filename):
    try:

        file = File1()
        file.name = filename
        file.save()
        wb2 = load_workbook(filename)
        sheet_names = wb2.get_sheet_names()  #
        ws = wb2.get_sheet_by_name(sheet_names[0])  # 读第一张表为：index[0]
        max_row = ws.max_row
        for i in range(3, max_row+1):
            L = []
            for j in ws[i]:
                L.append(j.value)
            temp = select_sysUser.objects.get(user_id__exact=L[1])
            year = temp.achieve_year
            subject = L[2]
            tea_no = L[1]
            major = select_sysUser.objects.get(user_id=tea_no)
            major_id = major.major_id

            introduction = L[3]
            if L[4]:
                subject_property ='研究论文'
            if L[5]:
                subject_property = '工程设计'
            if L[6]:
                subject_property = '项目报告'
            if L[7]:
                subject_property = '综述报告'
            if L[8]:
                subject_property = '其他'

            obj = Topic(

                    year=year,
                    subject= subject,
                    tea_no= tea_no,
                    introduction= introduction,
                    research_papers=L[4],
                    engineering_design=L[5],
                    project_report=L[6],
                    summary_report=L[7],
                    other=L[8],
                    other_introduction=L[9],
                    combine_actual=L[10],
                    company_name=L[11],
                    implementation_state=L[12],
                    english_title=L[13],
                    major_id = major_id,
                    subject_property=subject_property

            )

            obj.save()
            ID = obj.id

            obj1 = Topic1(

                year=year,
                subject=subject,
                tea_no=tea_no,
                introduction=introduction,
                research_papers=L[4],
                engineering_design=L[5],
                project_report=L[6],
                summary_report=L[7],
                other=L[8],
                other_introduction=L[9],
                combine_actual=L[10],
                company_name=L[11],
                implementation_state=L[12],
                english_title=L[13],
                major_id=major_id,
                subject_property=subject_property,
                topic=ID
            )

            obj1.save()


        return 1, '文件上传成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def getTeaToopicnum(limit, offset, institutionid,majorid,search,sort,order):
    try:

        institution_id = institutionid
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year

        cursor = connection.cursor()

        sql = """SELECT a.id,a.name,a.user_id,b.topic_num,c.major FROM app_select_sysuser a
              LEFT JOIN app_teacher_topic_num b ON a.user_id=b.tea_no
              LEFT JOIN app_major c ON a.major_id=c.id

              WHERE  (a.achieve_year= """ + str(achieve_year)
        sql = sql + " and a.institution_id=" + str(institution_id) + ")"+"and (a.role_id=4 or a.role_id=3)"
        if majorid !='' and majorid!='-1':
            sql = sql + ' AND a.major_id='+str(majorid)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR b.topic_num LIKE \'%' + search + \
                  '%\' OR c.major LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)

        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        print sql
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = """SELECT COUNT(*) FROM app_select_sysuser a
              LEFT JOIN app_teacher_topic_num b ON a.user_id=b.tea_no
              LEFT JOIN app_major c ON a.major_id=c.id

              WHERE  (a.achieve_year= """ + str(achieve_year)
        sql = sql + " and a.institution_id=" + str(institution_id) + ")"+"and (a.role_id=4 or a.role_id=3)"
        if majorid !='' and majorid!='-1':
            sql = sql + ' AND a.major_id='+str(majorid)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR b.topic_num LIKE \'%' + search + \
                  '%\' OR c.major LIKE \'%' + search + '%\')'

        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}

        return data
    except:
        return 0, '数据库错误'

def director_Numberbyid(eid):
    id = eid.replace('-', '')

    ID = select_sysUser.objects.get(id=id)
    teano = ID.user_id
    thisyear = Date_setting.objects.get(activation=1)
    achieveyear = thisyear.year
    try:
        cursor = connection.cursor()
        sql = """SELECT a.id,a.name,a.user_id,b.topic_num,c.major FROM app_select_sysuser a
                      LEFT JOIN app_teacher_topic_num b ON b.tea_no="""+str(teano)+" LEFT JOIN app_major c ON c.id=a.major_id"


        sql = sql + " WHERE  a.achieve_year="  + str(achieveyear)+" and a.id="+str(id)
        print sql


        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'




##韩佳琦


##许国斌 ， 教师  新增

def teacher_getTopicbyid(eid):
    id = eid.replace('-', '')
    try:
        cursor = connection.cursor()
        sql = '''SELECT * , b.major FROM app_topic1 a 
                 LEFT JOIN app_major b ON b.id=a.major_id  WHERE  a.id='''

        sql = sql + str(id)
        print sql
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def teacher_getTopicnumbyid(user_id):
    id = user_id.replace('-', '')
    thisyear = Date_setting.objects.get(activation=1)
    achieveyear = thisyear.year

    try:
        cursor = connection.cursor()
        sql = '''SELECT topic_num FROM app_teacher_topic_num  where tea_no=\''''

        sql = sql + str(id) +'\'' ' and year=' +'\''+ str(achieveyear)+'\''
        print (sql)
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def teacher_getTopicnum(user_id):
    id = user_id.replace('-', '')
    thisyear=Date_setting.objects.get(activation=1)
    achieveyear=thisyear.year
    try:
        cursor = connection.cursor()
        sql = '''SELECT * FROM app_topic1 where tea_no=\''''
        sql = sql + str(id) +'\'' ' and year=' +'\''+ str(achieveyear)+'\''
        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'




def getTeafile(userid,limit, offset):
    try:
        cursor = connection.cursor()
        sql = """SELECT SUBSTRING(file_position,8) AS filename,b.subject AS subjectname, file_position,b.tea_no from app_file a
           LEFT JOIN app_topic b ON a.subject = b.id
              WHERE b.tea_no="""
        sql = sql + str(userid)+' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        sql = """SELECT COUNT(*) FROM app_file a
                  LEFT JOIN app_topic b ON a.subject=b.id """
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        num=dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'] , 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'




def saveFile(filename,subject_id):
    try:
        file = File()
        file.file_position = filename
        file.subject = subject_id
        file.save()



        return 1, '文件上传成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info



def getTeaInfo(userid):
    try:
        dict = {}
        cursor = connection.cursor()
        cursor.execute("""
                  SELECT a.user_id,a.name,a.tel,b.sex,c.institution_name FROM app_select_sysuser a 
                  LEFT JOIN app_sex b  ON b.id=a.sex_id 
                  LEFT JOIN app_institution c ON a.institution_id=c.id
                  WHERE a.user_id=%s
                        """,[userid])
        dict['data'] = dictfetchall(cursor)


        return 1, dict
    except:
        return 0, '数据库错误'

def getTeaTopic(limit, offset, userid,search,sort,order):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.id,a.subject,a.introduction,a.year,b.selected_stu_no,b.submit,c.name,c.stu_class,c.user_id ,c.score,c.tel,d.sex,e.year,e.activation,f.major,g.stu_introduction FROM app_topic a
                 LEFT JOIN app_application_state b ON b.topic=a.id
                 LEFT JOIN app_select_sysuser c ON c.user_id=b.selected_stu_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 LEFT JOIN app_stu_introduction g ON g.stu_no=b.selected_stu_no
                 WHERE e.activation=1 AND a.tea_no='''+ str(userid)
        if search != '':
            sql = sql + ' AND (a.subject LIKE \'%' + search + \
                  '%\' OR d.sex LIKE \'%' + search + \
                  '%\' OR c.stu_class LIKE \'%' + search + \
                  '%\' OR c.score LIKE \'%' + search + \
                  '%\' OR c.tel LIKE \'%' + search + \
                  '%\' OR g.stu_introduction LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + \
                  '%\' OR c.name LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)
        sql = sql  + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        print sql
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = '''SELECT COUNT(*) FROM app_topic a
                 LEFT JOIN app_application_state b ON b.topic=a.id
                 LEFT JOIN app_select_sysuser c ON c.user_id=b.selected_stu_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 LEFT JOIN app_stu_introduction g ON g.stu_no=b.selected_stu_no
                 WHERE e.activation=1 AND a.tea_no='''+ str(userid)
        if search != '':
            sql = sql + ' AND (a.subject LIKE \'%' + search + \
                  '%\' OR d.sex LIKE \'%' + search + \
                  '%\' OR c.stu_class LIKE \'%' + search + \
                  '%\' OR c.score LIKE \'%' + search + \
                  '%\' OR c.tel LIKE \'%' + search + \
                  '%\' OR g.stu_introduction LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + \
                  '%\' OR c.name LIKE \'%' + search + '%\')'
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'

def getTeaTopic1(limit, offset, userid,search,sort,order):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.id,a.subject,a.introduction,a.year,b.selected_stu_no,b.submit,c.name,c.stu_class,c.user_id ,d.sex,e.year,e.activation,f.major FROM app_topic1 a
                 LEFT JOIN app_application_state b ON b.topic=a.id
                 LEFT JOIN app_select_sysuser c ON c.user_id=b.selected_stu_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 WHERE e.activation=1 AND a.tea_no='''+ str(userid)
        if search != '':
            sql = sql + ' AND (a.subject LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + \
                  '%\' OR f.major LIKE \'%' + search + '%\')'
        if sort != '':
            sql = sql + ' ORDER BY ' + str(sort) + ' ' + str(order)
        sql = sql  + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)

        cursor.execute(sql)
        select_out = dictfetchall(cursor)


        sql = '''SELECT COUNT(*) FROM app_topic1 a
                 LEFT JOIN app_application_state b ON b.topic=a.id
                 LEFT JOIN app_select_sysuser c ON c.user_id=b.selected_stu_no
                 LEFT JOIN  app_sex d ON c.sex_id=d.id
                 LEFT JOIN  app_date_setting e ON a.year=e.year
                 LEFT JOIN app_major f ON f.id=a.major_id
                 WHERE e.activation=1 AND a.tea_no='''+ str(userid)
        if search != '':
            sql = sql + ' AND (a.subject LIKE \'%' + search + \
                  '%\' OR a.introduction LIKE \'%' + search + \
                  '%\' OR f.major LIKE \'%' + search + '%\')'

        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}

        return data
    except:
        return 0, '数据库错误'

def addTopic(
                userid,
                subject,
                introduction,
                subject_property_id,
                other_introduction,
                combine_actual,
                company_name,
                implementation_state,
                english_title
            ):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    teamajor = select_sysUser.objects.get(user_id=userid)
    majorid = teamajor.major_id
    try:

        if combine_actual=='是':
            combine_actual='√'
        if subject_property_id =='研究论文':

            obj = Topic1(
                tea_no=userid,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                research_papers='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id = majorid
            )
            obj.save()
            return 1, '添加成功'

        if subject_property_id == '工程设计':

            obj = Topic1(
                tea_no=userid,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                engineering_design='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id = majorid
            )
            obj.save()
            return 1, '添加成功'

        if subject_property_id == '项目报告':

            obj = Topic1(
                tea_no=userid,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                project_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=majorid

            )
            obj.save()
            return 1, '添加成功'

        if subject_property_id == '综述报告':

            obj = Topic1(
                tea_no=userid,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                summary_report='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=majorid
            )
            obj.save()
            return 1, '添加成功'

        if subject_property_id == '其他':

            obj = Topic1(
                tea_no=userid,
                subject=subject,
                year=achieve_year,
                introduction=introduction,
                subject_property=subject_property_id,
                other='√',
                other_introduction=other_introduction,
                combine_actual=combine_actual,
                company_name=company_name,
                implementation_state=implementation_state,
                english_title=english_title,
                major_id=majorid
            )
            obj.save()
            return 1, '添加成功'

    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def updateTopic(
        id,
        userid,
        subject,
        introduction,
        subject_property_id,
        other_introduction,
        combine_actual,
        company_name,
        implementation_state,
        english_title
):
    try:
        teamajor = select_sysUser.objects.get(user_id=userid)
        majorid = teamajor.major_id
        if combine_actual=='是':
            combine_actual='√'
        count = Topic1.objects.filter(id=id)
        if len(count) == 0:
            return 0, '题目不存在'
        temp = Topic1.objects.get(id=id)
        if temp.topic!='':
            return 0,'提交后不能修改，请联系教学所长！'
        if subject_property_id=='研究论文':
           object = Topic1.objects.get(id=id)
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = '√'
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.major_id = majorid
           object.save()
           return 1, '更新成功'

        if subject_property_id=='工程设计':
           object = Topic1.objects.get(id=id)
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = '√'
           object.project_report = ''
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.major_id = majorid
           object.save()
           return 1, '更新成功'

        if subject_property_id=='项目报告':
           object = Topic1.objects.get(id=id)
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = '√'
           object.summary_report = ''
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.major_id = majorid
           object.save()
           return 1, '更新成功'

        if subject_property_id=='综述报告':
           object = Topic1.objects.get(id=id)
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = '√'
           object.other = ''
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.major_id = majorid
           object.save()
           return 1, '更新成功'

        if subject_property_id=='其他':
           object = Topic1.objects.get(id=id)
           object.subject = subject
           object.introduction = introduction
           object.subject_property = subject_property_id
           object.research_papers = ''
           object.engineering_design = ''
           object.project_report = ''
           object.summary_report = ''
           object.other = '√'
           object.other_introduction = other_introduction
           object.combine_actual = combine_actual
           object.company_name = company_name
           object.implementation_state = implementation_state
           object.english_title = english_title
           object.major_id = majorid
           object.save()
           return 1, '更新成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def getStudent1(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=1'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStudent2(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=2'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStudent3(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=3'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStuinfo(student):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.stu_class,b.sex,a.score,c.stu_introduction FROM app_select_sysuser a
                 LEFT JOIN  app_sex b ON a.sex_id=b.id
                 LEFT JOIN  app_stu_introduction c ON c.stu_no='''+str(student)
        sql = sql +  " WHERE a.user_id="+str(student)
        print sql
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getState(subject_id):
    try:
        cursor = connection.cursor()
        sql='''SELECT a.id FROM app_topic a 
               WHERE a.id='''+"'"+str(subject_id)+"'"
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def addStudent(
        id,
        topic,
        selected_stu_no
    ):

    try:

        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        teacher = Topic.objects.get(id=topic)
        teano = teacher.tea_no
        TOPIC = Application_state.objects.filter(topic=topic)
        cursor = connection.cursor()
        sql='''select * from app_application_state a 
                left join app_topic b on b.id=a.topic where b.year='''+achieve_year+" and b.tea_no="+teano
        cursor.execute(sql)

        select_out = dictfetchall(cursor)
        sql = '''SELECT topic_num FROM app_teacher_topic_num

                              WHERE tea_no=''' + teano + " AND year=" + achieve_year
        cursor.execute(sql)
        aa = dictfetchall(cursor)
        if len(aa)==0:
            return 1, '教学所长尚未给您设定题目数量，还不能选择学生！'
        num = aa[0]

        if len(TOPIC)!=0:
            return 1,'您已为该题目选择完学生，不能更改！'
        elif len(select_out)>=int(num['topic_num']):

            return 1,'您已经选择了'+str(num['topic_num'])+'名学生，不能继续选择了！'
        else:
            obj = Application_state(
                id=id,
                topic=topic,
                selected_stu_no=selected_stu_no,
                submit=1)
            obj.save()

            return 1, '保存成功'


    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def getTeachertopics(usr_id):
    try:
        cursor = connection.cursor()
        sql ='''SELECT a.id,a.year,a.subject,a.tea_no,b.year,b.activation FROM app_topic a
                 LEFT JOIN  app_date_setting b ON a.year=b.year
                 WHERE b.activation=1 AND a.tea_no='''+str(usr_id)
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def teacher_getDate(tea_no):
    try:
        dict = {}
        cursor = connection.cursor()
        cursor.execute("""
                        SELECT * FROM app_date_setting a
                        WHERE a.activation=1 
                                    """)
        dict['date'] = dictfetchall(cursor)

        return 1, dict
    except:
        return 0, '数据库错误'





##许国斌



##王恩瑞 ， 教办管理员   新增

def getSelectstulast(limit, offset):

    try:
        cursor = connection.cursor()
        '''
        sql = """select tea_no,
            SUBSTRING_INDEX(k,',',1) as name1,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) as name2,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) as name3,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) as name4
            from (
            select tea_no,group_concat(subject) as k  
            from app_topic
            group by tea_no
            ) t """
        '''
        sql='''SELECT name,user_id,sex,stu_class,stu_introduction,institution_name
        FROM app_select_sysuser a
        LEFT JOIN app_sex b ON a.sex_id = b.id
        LEFT JOIN app_institution c ON a.institution_id = c.id
        LEFT JOIN app_stu_introduction e ON a.user_id = e.stu_no
        WHERE role_id = 2'''
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        data = {'total': cursor.execute('''SELECT COUNT(*) '''), 'row': select_out}
        return data
    except:
        return 0, '数据库错误'

def getDowndata(limit, offset, search,major_id,class_id,institution_id, teacher_id):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    try:
        cursor = connection.cursor()

        sql='''SELECT a.id,a.name,a.user_id,b.sex,a.stu_class,e.major,c.institution_name,g.name AS Bname,f.subject
        FROM app_select_sysuser a
        LEFT JOIN app_sex b ON a.sex_id = b.id        
        LEFT JOIN app_application_state d ON a.user_id = d.selected_stu_no
        LEFT JOIN app_major e ON a.major_id = e.id
        LEFT JOIN app_topic f ON d.topic = f.id
        LEFT JOIN app_select_sysuser AS g ON f.tea_no = g.user_id
        LEFT JOIN app_institution c ON g.institution_id = c.id
        LEFT JOIN app_class AS h ON a.stu_class = h.class_name
        WHERE a.role_id = 1 and a.achieve_year = '''+str(achieve_year)

        if institution_id != '' and institution_id != '-1':
            sql = sql + ' AND g.institution_id = ' + str(institution_id)
        if major_id != '' and major_id != '-1':
            sql = sql + ' AND a.major_id = ' + str(major_id)
        if teacher_id != '' and teacher_id != '-1':
            sql = sql + ' AND g.id = ' + str(teacher_id)
        if class_id != '' and class_id != '-1':
            sql = sql + ' AND h.id = ' + str(class_id)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR a.user_id LIKE \'%' + search + \
                  '%\' OR e.major LIKE \'%' + search + \
                  '%\' OR g.name LIKE \'%' + search + \
                  '%\' OR institution_name LIKE \'%' + search + '%\')'

        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = '''SELECT COUNT(*)FROM app_select_sysuser a
                LEFT JOIN app_sex b ON a.sex_id = b.id        
                LEFT JOIN app_application_state d ON a.user_id = d.selected_stu_no
                LEFT JOIN app_major e ON a.major_id = e.id
                LEFT JOIN app_topic f ON d.topic = f.id
                LEFT JOIN app_select_sysuser AS g ON f.tea_no = g.user_id
                LEFT JOIN app_institution c ON g.institution_id = c.id
                LEFT JOIN app_class AS h ON a.stu_class = h.class_name
                WHERE a.role_id = 1 and a.achieve_year = ''' + str(achieve_year)

        if institution_id != '' and institution_id != '-1':
            sql = sql + ' AND g.institution_id = ' + str(institution_id)
        if major_id != '' and major_id != '-1':
            sql = sql + ' AND a.major_id = ' + str(major_id)
        if teacher_id != '' and teacher_id != '-1':
            sql = sql + ' AND g.id = ' + str(teacher_id)
        if class_id != '' and class_id != '-1':
            sql = sql + ' AND h.id = ' + str(class_id)
        if search != '':
            sql = sql + ' AND (a.name LIKE \'%' + search + \
                  '%\' OR a.user_id LIKE \'%' + search + \
                  '%\' OR g.name LIKE \'%' + search + \
                  '%\' OR e.major LIKE \'%' + search + \
                  '%\' OR institution_name LIKE \'%' + search + '%\')'
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data

    except:
        return 0, '数据库错误'

def getSelectbystu(limit, offset):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    try:
        cursor = connection.cursor()
        sql='''SELECT a.name,a.user_id,b.sex,a.stu_class,c.institution_name,g.name AS Bname,f.subject
        FROM app_select_sysuser a
        LEFT JOIN app_sex b ON a.sex_id = b.id
        LEFT JOIN app_institution c ON a.institution_id = c.id
        LEFT JOIN app_application_state d ON a.user_id = d.selected_stu_no
        LEFT JOIN app_topic f ON d.topic = f.id
        LEFT JOIN app_select_sysuser AS g ON f.tea_no = g.user_id
        WHERE a.role_id = 1 and a.achieve_year = '''+str(achieve_year)
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = '''SELECT COUNT(*) FROM app_select_sysuser a
        LEFT JOIN app_sex b ON a.sex_id = b.id
        LEFT JOIN app_institution c ON a.institution_id = c.id
        LEFT JOIN app_application_state d ON a.user_id = d.selected_stu_no
        LEFT JOIN app_topic f ON d.topic = f.id
        LEFT JOIN app_select_sysuser AS g ON f.tea_no = g.user_id
       WHERE a.role_id = 1 and a.achieve_year = '''+str(achieve_year)

        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data

    except:
        return 0, '数据库错误'


def getSelectbysub(limit, offset):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year
    try:
        cursor = connection.cursor()
        sql = '''select t.id,a.name,subject
            FROM app_topic t
            LEFT JOIN app_select_sysuser a ON t.tea_no = a.user_id

            WHERE t.id not in (select app_topic.id FROM app_topic
            INNER JOIN app_application_state b ON app_topic.id = b.topic) and t.year = '''+str(achieve_year)
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        sql = '''SELECT COUNT(*) FROM app_topic t
            LEFT JOIN app_select_sysuser a ON t.tea_no = a.user_id

            WHERE t.id not in (select app_topic.id FROM app_topic
            INNER JOIN app_application_state b ON app_topic.id = b.topic) and t.year = '''+str(achieve_year)

        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data

    except:
        return 0, '数据库错误'

def getSelectbytea(limit, offset):
    achieveyear = Date_setting.objects.get(activation=1)
    achieve_year = achieveyear.year

    try:
        cursor = connection.cursor()
        sql = '''select a.name,c.name AS name1,e.name AS name2,g.name AS name3,i.name AS name4,
            j.subject AS subject1,k.subject AS subject2,l.subject AS subject3,m.subject AS subject4,
            SUBSTRING_INDEX(k,',',1) as id1,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) as id2,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) as id3,
            SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) as id4
            from app_select_sysuser a
            LEFT JOIN (
            select id,year,tea_no,group_concat(id) as k  
            from app_topic
            where(year = '''+str(achieve_year)+')' '''
            group by tea_no
            ) t ON a.user_id = t.tea_no
            
            LEFT JOIN app_application_state b ON SUBSTRING_INDEX(k,',',1) = b.topic
            LEFT JOIN app_select_sysuser AS c ON b.selected_stu_no = c.user_id
            LEFT JOIN app_topic AS j ON SUBSTRING_INDEX(k,',',1) = j.id
            LEFT JOIN app_application_state AS d ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) = d.topic
            LEFT JOIN app_select_sysuser AS e ON d.selected_stu_no = e.user_id
            LEFT JOIN app_topic AS k ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) = k.id
            LEFT JOIN app_application_state AS f ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) = f.topic
            LEFT JOIN app_select_sysuser AS g ON f.selected_stu_no = g.user_id
            LEFT JOIN app_topic AS l ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) = l.id
            LEFT JOIN app_application_state AS h ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) = h.topic
            LEFT JOIN app_select_sysuser AS i ON h.selected_stu_no = i.user_id
            LEFT JOIN app_topic AS m ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) = m.id
            WHERE (a.role_id = 4 or a.role_id = 3)'''

        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        for br in select_out:
            if br['subject4'] == br['subject3']:
                br['subject4'] = None
                br['name4'] = None
                if br['subject3'] == br['subject2']:
                    br['subject3'] = None
                    br['name3'] = None
                    if br['subject2'] == br['subject1']:
                        br['subject2'] = None
                        br['name2'] = None

        sql = '''SELECT COUNT(*) from app_select_sysuser a
        
            LEFT JOIN (
            select id,year,tea_no,group_concat(id) as k  
            from app_topic
            where(year = '''+str(achieve_year)+')' '''
            group by tea_no
            ) t ON a.user_id = t.tea_no
            
            LEFT JOIN app_application_state b ON SUBSTRING_INDEX(k,',',1) = b.topic
            LEFT JOIN app_select_sysuser AS c ON b.selected_stu_no = c.user_id
            LEFT JOIN app_topic AS j ON SUBSTRING_INDEX(k,',',1) = j.id
            LEFT JOIN app_application_state AS d ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) = d.topic
            LEFT JOIN app_select_sysuser AS e ON d.selected_stu_no = e.user_id
            LEFT JOIN app_topic AS k ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',2),',',-1) = k.id
            LEFT JOIN app_application_state AS f ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) = f.topic
            LEFT JOIN app_select_sysuser AS g ON f.selected_stu_no = g.user_id
            LEFT JOIN app_topic AS l ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',3),',',-1) = l.id
            LEFT JOIN app_application_state AS h ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) = h.topic
            LEFT JOIN app_select_sysuser AS i ON h.selected_stu_no = i.user_id
            LEFT JOIN app_topic AS m ON SUBSTRING_INDEX(SUBSTRING_INDEX(k,',',4),',',-1) = m.id
            WHERE (a.role_id = 4 or a.role_id = 3)'''

        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data

    except:
        return 0, '数据库错误'


def getClass():
    try:
        cursor = connection.cursor()
        cursor.execute(""" SELECT id, class_name FROM app_class """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getTeacher(institution_id):
    try:
        cursor = connection.cursor()

        print institution_id
        sql = '''select id,name from app_select_sysuser where (role_id = 4 or role_id=3)'''


        if institution_id != '' and institution_id != '-1':
            sql = sql + ' AND institution_id = ' + str(institution_id)

        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def getInstitution():

    try:
        cursor = connection.cursor()
        cursor.execute(""" select id, institution_name from app_institution """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'



##王恩瑞

def supervisor_getDate():
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_date_setting where activation=1""")
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'


def director_getAlltopicnum(userid):
    try:

        user = select_sysUser.objects.get(user_id=userid)
        institutionid = user.institution_id
        thisyear = Date_setting.objects.get(activation=1)
        achieveyear = thisyear.year
        cursor = connection.cursor()
        sql = """ select * from app_select_sysuser a
                  where role_id=1 and institution_id = 
              """
        sql = sql + str(institutionid)+ ' and achieve_year='+str(achieveyear)

        cursor.execute(sql)

        return 1, dictfetchall(cursor)

    except:
            return 0, '数据库错误'


def director_getDividedtopicnum(userid):
    try:
        user = select_sysUser.objects.get(user_id=userid)
        institutionid = user.institution_id
        thisyear = Date_setting.objects.get(activation=1)
        achieveyear = thisyear.year
        cursor = connection.cursor()
        sql = """ select * from app_teacher_topic_num a
                  LEFT JOIN app_select_sysuser b ON  a.tea_no=b.user_id 
                  where institution_id = 
              """
        sql = sql + str(institutionid) + ' and a.year='+ str(achieveyear)

        cursor.execute(sql)
        return 1, dictfetchall(cursor)

    except:
            return 0, '数据库错误'

def director_getResttopicnum(userid):
    try:

        user = select_sysUser.objects.get(user_id=userid)
        institutionid = user.institution_id
        thisyear = Date_setting.objects.get(activation=1)
        achieveyear = thisyear.year
        cursor = connection.cursor()
        sql = """ SELECT a.name,a.role_id,a.user_id,a.achieve_year,b.tea_no,b.topic_num,b.year,a.institution_id FROM app_select_sysuser a 
        LEFT JOIN app_teacher_topic_num b ON a.user_id=b.tea_no 
        WHERE a.institution_id="""

        sql = sql + str(institutionid) + " and (a.achieve_year=" + str(achieveyear)+" or b.year="+ str(achieveyear)+")"

        cursor.execute(sql)


        return 1, dictfetchall(cursor)

    except:
            return 0, '数据库错误'



##11.11 ltl
def director_getMajor(institution_id):
    try:
        cursor = connection.cursor()

        sql = "SELECT  * FROM app_major"


        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStudent4(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=4'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStudent5(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=5'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def getStudent6(subject_id):

    try:
        cursor = connection.cursor()
        sql = '''SELECT  a.user_id,a.name,b.submit,c.volunteer_topic_no_id FROM app_select_sysuser a
                 LEFT JOIN app_application_state b ON b.selected_stu_no=a.user_id
                 LEFT JOIN app_application c ON a.user_id=c.stu_no
                 WHERE (a.user_id NOT IN(SELECT b.selected_stu_no FROM app_application_state b) AND c.volunteer_topic_no_id='''+str(subject_id)+')'+'''AND c.volunteer_no=6'''
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

## 11.11 zwc

def supervisor_getMajor():
    try:
        cursor = connection.cursor()
        cursor.execute(""" select * from app_major """)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'



##11.12 zwc

def supervisor_getTopicnum(limit, offset,search):
#每个LEFT JOIN(）里加了一个年份控制，不然的话，数据库里有两个年份的话会显示错误
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year

        cursor = connection.cursor()
        sql = """SELECT a.id,b.institution_name,a.min,topic_num,unselect_num, select_num FROM app_institution_topic_num a
                 LEFT JOIN app_institution b ON b.id = a.institution_id
                 LEFT JOIN( SELECT c.id AS cid,COUNT(*) topic_num FROM app_institution_topic_num c                 
                      LEFT JOIN app_select_sysuser d ON d.institution_id = c.institution_id
                      LEFT JOIN app_topic e ON e.tea_no = d.user_id WHERE e.tea_no!='' AND c.year="""+str(achieve_year)+"""               
                      GROUP BY c.institution_id
                 ) T1 ON T1.cid = a.id
                 LEFT JOIN( SELECT f.id AS fid,COUNT(*) select_num FROM app_institution_topic_num f                
                      LEFT JOIN app_select_sysuser g ON g.institution_id = f.institution_id
                      LEFT JOIN app_topic h ON h.tea_no = g.user_id
                      LEFT JOIN app_application_state i ON i.topic = h.id WHERE i.id!='' AND f.year="""+str(achieve_year)+"""              
                      GROUP BY f.institution_id
                 ) T2 ON T2.fid = a.id
                 LEFT JOIN( SELECT j.id AS jid,COUNT(*) unselect_num FROM app_institution_topic_num j               
                      LEFT JOIN app_select_sysuser k ON k.institution_id = j.institution_id
                      LEFT JOIN app_topic l ON l.tea_no = k.user_id                 
                      WHERE l.id!=''AND j.year="""+str(achieve_year)+""" AND l.id NOT IN (SELECT topic FROM app_application_state)               
                      GROUP BY j.institution_id
                 ) T3 ON T3.jid = a.id              
                 WHERE a.year = """ + str(achieve_year)
        print sql
        sql = sql + ' LIMIT ' + str(limit) + ' OFFSET ' + str(offset)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        print 'ooo'
        print select_out

        sql = """SELECT COUNT(*) FROM app_institution_topic_num WHERE year = """ + str(achieve_year)
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]
        data = {'total': num['COUNT(*)'], 'rows': select_out}
        return data
    except:
        return 0, '数据库错误'

def supervisor_setTopicnum(
        institution_id,
        topic_min
       ):
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        institution = institution_topic_num.objects.filter(institution_id=institution_id)
        current_inst = institution.filter(year=achieve_year)
        if len(current_inst) > 0:
            return 0, '该研究所已设置'
        else:
            obj = institution_topic_num(
                institution_id = institution_id,
                min = topic_min,
                year = achieve_year
            )
            obj.save()

        return 1, '设置成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info





##11.18 ltl

def director_Get_minmax(institution_id):

    try:

        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        cursor = connection.cursor()

        #应分配题目数
        sql ='''select min from app_institution_topic_num 
                where year='''+str(achieve_year)+" and institution_id="+str(institution_id)
        cursor.execute(sql)
        a = dictfetchall(cursor)

        # 已分配题目数
        sql = '''select a.topic_num from app_teacher_topic_num a 
                          left join app_select_sysuser b on b.user_id = a.tea_no
                          where a.year=''' + str(achieve_year) + " and b.institution_id=" + str(institution_id)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)

        num = 0
        for item in select_out:

            if item['topic_num'] == '' or item['topic_num'] == None:
                num += 0
            else:
                num += int(item['topic_num'])

        a[0].setdefault('topic_num',num)



        # 剩余应分配题目数
        rest_num=0
        b=int(a[0]['min'])-int(a[0]['topic_num'])

        if b>0:
            rest_num=b
        else:
            rest_num='已满足要求'
        a[0].setdefault('rest_num', rest_num)

        return 1, a
    except:
        return 0, '数据库错误'


def director_getTopicnum(institution_id):

    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year

        cursor = connection.cursor()
        """
        sql = '''select min from app_institution_topic_num 
                  where institution_id='''+str(institution_id)+" and year="+str(achieve_year)
        cursor.execute(sql)
        a = dictfetchall(cursor)

        #已分配题目数
        sql = '''select a.topic_num from app_teacher_topic_num a 
                  left join app_select_sysuser b on b.user_id = a.tea_no
                  where a.year='''+str(achieve_year)+" and b.institution_id="+str(institution_id)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        num=0
        for item in select_out:
            if item['topic_num']=='' or item['topic_num']==None:
                num+=0
            num+=int(item['topic_num'])

        #已填报题目数
        sql = '''select a.tea_no from app_topic a 
                  left join app_select_sysuser b on b.user_id = a.tea_no
                  where a.year='''+str(achieve_year)+" and b.institution_id="+str(institution_id)
        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        a[0].setdefault('topic_num', len(select_out))
        """
        sql = """SELECT a.id,b.institution_name,max,min,topic_num,select_num ,unselect_num, select_num FROM app_institution_topic_num a
                         LEFT JOIN app_institution b ON b.id = a.institution_id
                         LEFT JOIN( SELECT c.id AS cid,COUNT(*) topic_num FROM app_institution_topic_num c                 
                              LEFT JOIN app_select_sysuser d ON d.institution_id = c.institution_id
                              LEFT JOIN app_topic e ON e.tea_no = d.user_id WHERE e.tea_no!='' AND c.year=""" + str(
            achieve_year) + """               
                              GROUP BY c.institution_id
                         ) T1 ON T1.cid = a.id
                         LEFT JOIN( SELECT f.id AS fid,COUNT(*) select_num FROM app_institution_topic_num f                
                              LEFT JOIN app_select_sysuser g ON g.institution_id = f.institution_id
                              LEFT JOIN app_topic h ON h.tea_no = g.user_id
                              LEFT JOIN app_application_state i ON i.topic = h.id WHERE i.id!='' AND f.year=""" + str(
            achieve_year) + """              
                              GROUP BY f.institution_id
                         ) T2 ON T2.fid = a.id
                         LEFT JOIN( SELECT j.id AS jid,COUNT(*) unselect_num FROM app_institution_topic_num j               
                              LEFT JOIN app_select_sysuser k ON k.institution_id = j.institution_id
                              LEFT JOIN app_topic l ON l.tea_no = k.user_id                 
                              WHERE l.id!=''AND j.year=""" + str(achieve_year) + """ AND l.id NOT IN (SELECT topic FROM app_application_state)               
                              GROUP BY j.institution_id
                         ) T3 ON T3.jid = a.id              
                         WHERE a.year = """ + str(achieve_year)+" and b.id="+str(institution_id)
        cursor.execute(sql)

        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def Giveup1(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"1"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "1)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                1) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'

def Giveup2(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"2"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "2)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                2) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'

def Giveup3(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"3"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "3)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                3) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'

def Giveup4(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"4"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "4)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                4) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'

def Giveup5(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"5"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "5)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                5) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'

def Giveup6(subject_id):
    try:
        temp = Application_state.objects.filter(topic=subject_id)
        if len(temp) != 0:
            return 1, '选择完学生后不可弃权！'
        cursor = connection.cursor()
        sql = '''SELECT * FROM `app_giveup_topic` WHERE topic='''+subject_id + " and day_num="+"6"
        cursor.execute(sql)
        select = dictfetchall(cursor)
        if len(select) != 0:
            return 1, '该题目已经弃权，请勿重复操作！'
        else:
            cursor = connection.cursor()
            sql = '''INSERT INTO `app_giveup_topic`(`id`, `topic`, `day_num`) VALUES(NULL, '''+subject_id+","+ "6)"
            cursor.execute(sql)
            cursor = connection.cursor()
            sql = """DELETE FROM app_application WHERE volunteer_no=""" + str(
                6) + " and volunteer_topic_no_id=" + subject_id
            cursor.execute(sql)
            return 1, '弃权成功'
    except:
        return 0, '数据库错误'


##11.18 zwc

def supervisor_getTopicNumbyid(eid):
    id = eid.replace('-', '')

    try:
        print "daozhele111"
        cursor = connection.cursor()
        sql = '''SELECT * FROM app_institution_topic_num WHERE id='''+ str(id)
        cursor.execute(sql)
        print "daozhele222"
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'




def supervisor_updateTopicNum(
        id,
        institution_id,
        topic_min
):
    try:
        count = institution_topic_num.objects.filter(id=id)
        if len(count) == 0:
            return 0, '研究所不存在'
        else:
            object = institution_topic_num.objects.get(id=id)

            object.institution_id = institution_id
            object.min = topic_min

            object.save()
            return 1, '更新成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

## 11.18 wer
def student_getStuInfo(userid):
    try:
        dict = {}
        cursor = connection.cursor()
        cursor.execute("""
                  SELECT a.user_id,a.name,a.stu_class,a.tel,b.sex  
                  FROM app_select_sysuser a 
                  LEFT JOIN app_sex b  ON b.id=a.sex_id 
                  WHERE a.user_id=%s
                        """,[userid])
        dict['data'] = dictfetchall(cursor)

        cursor.execute("""
                  SELECT b.major,b.id 
                  FROM app_select_sysuser a 
                  LEFT JOIN app_major b ON a.major_id=b.id  
                  WHERE a.user_id = %s
                         """, [userid])
        dict['major'] = dictfetchall(cursor)

        return 1, dict

    except:
        return 0, '数据库错误'


def student_addIntruduction(userid,stu_introduction):

    try:
        introduction = Stu_introduction.objects.filter(stu_no=userid)
        if len(introduction)==0:
            obj = Stu_introduction(
                stu_no=userid,
                stu_introduction=stu_introduction)
            obj.save()
            return 1, '保存成功'
        else:
            object = Stu_introduction.objects.get(stu_no=userid)
            object.stu_introduction = stu_introduction
            object.save()
            return 1, '保存成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

def student_getIntroduction(stu_no):
    try:
        introduction = Stu_introduction.objects.get(stu_no = stu_no)
        return 1, introduction.stu_introduction
    except:
        return 0, ''





## 11.22 wer
def stu_SubmitApplication(submit,id1,id2,id3,id4,id5,id6):

    try:
        obj1 = Application.objects.get(id=id1)
        obj1.submit = submit
        obj1.save()
        obj2 = Application.objects.get(id=id2)
        obj2.submit = submit
        obj2.save()
        obj3 = Application.objects.get(id=id3)
        obj3.submit = submit
        obj3.save()
        obj4 = Application.objects.get(id=id4)
        obj4.submit = submit
        obj4.save()
        obj5 = Application.objects.get(id=id5)
        obj5.submit = submit
        obj5.save()
        obj6 = Application.objects.get(id=id6)
        obj6.submit = submit
        obj6.save()
        return 1, '保存成功'

    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

## 11.27 ltl
def Teacher_submit_topic(userid):
    try:
        print 'mei1'
        achieveyear = Date_setting.objects.get(activation=1)
        year = achieveyear.year
        cursor = connection.cursor()
        sql = '''SELECT * FROM app_topic1 
              
              WHERE tea_no='''+userid +" AND year="+year

        cursor.execute(sql)
        select_out = dictfetchall(cursor)
        submit_num = len(select_out)
        sql = '''SELECT topic_num FROM app_teacher_topic_num

                      WHERE tea_no=''' + userid + " AND year=" + year
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]

        if int(submit_num) < int(num['topic_num']):
            return 0,'您提交的题目数量不足'+num['topic_num']+'个，请继续上传题目！'

        for a in select_out:

            if a['topic']=='':
                #提交
                print 'haha'
                obj = Topic(
                    tea_no=a['tea_no'],
                    subject=a['subject'],
                    year=a['year'],
                    introduction=a['introduction'],
                    subject_property=a['subject_property'],
                    research_papers=a['research_papers'],
                    other_introduction=a['other_introduction'],
                    combine_actual=a['combine_actual'],
                    company_name=a['company_name'],
                    implementation_state=a['implementation_state'],
                    english_title=a['english_title'],
                    major_id=a['major_id'],
                    summary_report=a['summary_report'],
                    other=a['other'],
                    engineering_design=a['engineering_design'],
                    project_report=a['project_report']
                )
                obj.save()
                topic_id = obj.id
                TOPIC = Topic1.objects.get(id=a['id'])
                TOPIC.topic = topic_id
                TOPIC.save()

            else:
                #修改

                TOPIC = Topic.objects.get(id=a['topic'])

                TOPIC.tea_no = a['tea_no']
                TOPIC.subject = a['subject']
                TOPIC.year = a['year']
                TOPIC.introduction = a['introduction']
                TOPIC.subject_property = a['subject_property']
                TOPIC.research_papers = a['research_papers']
                TOPIC.other_introduction = a['other_introduction']
                TOPIC.combine_actual = a['combine_actual']
                TOPIC.company_name = a['company_name']
                TOPIC.implementation_state = a['implementation_state']
                TOPIC.english_title = a['english_title']
                TOPIC.major_id = a['major_id']
                TOPIC.summary_report = a['summary_report']
                TOPIC.other = a['other']
                TOPIC.engineering_design = a['engineering_design']
                TOPIC.project_report = a['project_report']
                TOPIC.save()


        return 1, '提交成功，您目前有'+str(submit_num)+'个题目！'

    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info


def Teacher_whether_submit(userid):
    try:
        print 'you1'
        achieveyear = Date_setting.objects.get(activation=1)
        year = achieveyear.year
        cursor = connection.cursor()
        sql = '''SELECT * FROM app_topic1 

              WHERE tea_no=''' + userid + " AND year=" + year

        cursor.execute(sql)
        select_out = dictfetchall(cursor)


        sql = '''SELECT topic_num FROM app_teacher_topic_num

                              WHERE tea_no=''' + userid + " AND year=" + year
        cursor.execute(sql)
        num = dictfetchall(cursor)[0]

        if int(len(select_out)) < int(num['topic_num']):
            return 0, '您提交的题目数量不足' + num['topic_num'] + '个，请继续上传题目！'

        for a in select_out:
            if a['topic']=='':
                return 1, '没提交过'

        return 1,'提交过了'



    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info

#12.05 ltl
def director_getTopicbyid(eid):
    id = eid.replace('-', '')
    try:
        cursor = connection.cursor()
        sql = '''SELECT * , b.major FROM app_topic a 
                 LEFT JOIN app_major b ON b.id=a.major_id  WHERE  a.id='''

        sql = sql + str(id)

        cursor.execute(sql)
        return dictfetchall(cursor)
    except:
        return 0, '数据库错误'

##12.05 zwc
def supervisor_getThisMajor(institution_id):
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        cursor = connection.cursor()
        sql = ''' SELECT a.major_id,b.major FROM app_select_sysuser a 
                  LEFT JOIN app_major b ON b.id = a.major_id 
                  WHERE institution_id = ''' + str(institution_id) + ''' 
                  AND achieve_year = ''' + str(achieve_year) + ''' GROUP BY major_id '''
        print sql
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

## 12.06 ltl
def director_getThisMajor(institution_id):
    try:
        achieveyear = Date_setting.objects.get(activation=1)
        achieve_year = achieveyear.year
        cursor = connection.cursor()
        sql = ''' SELECT a.major_id,b.major FROM app_select_sysuser a 
                  LEFT JOIN app_major b ON b.id = a.major_id 
                  WHERE institution_id = ''' + str(institution_id) + ''' 
                  AND achieve_year = ''' + str(achieve_year) + ''' GROUP BY major_id '''
        print sql
        cursor.execute(sql)
        return 1, dictfetchall(cursor)
    except:
        return 0, '数据库错误'

def teacher_delTopic(id):
    try:
        cursor = connection.cursor()
        sql = '''DELETE FROM app_topic1 WHERE id =''' + id
        cursor.execute(sql)
        return 1, '删除成功'
    except:
        import sys
        info = "%s || %s" % (sys.exc_info()[0], sys.exc_info()[1])
        return 0, info