# coding=utf-8

from DBoperater import uuidToString, dictfetchall, uuid, toUUID
from django.db import connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render_to_response
import django
django.setup()
import json
import os
import datetime
import sys
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Alignment

import time

message = '成功创建工程: '
print message


message = str(message)
print message

exit()


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def clear_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style=None, color="00FFFFFF")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border



def test():
    f_name = 'templates/blast_record_210000.xlsx'
    wb = load_workbook(filename=f_name)
    ws = wb.active

    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM')
    date_style.alignment = Alignment(horizontal='general', vertical='center')
    wrap_alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True)

    set_border(ws, "A3:S8")
    set_border(ws, "R3:R4")
    set_border(ws, "B11:R11")
    set_border(ws, "B13:R13")
    set_border(ws, "B15:R15")
    set_border(ws, "B17:R17")
    set_border(ws, "B10:N18")
    set_border(ws, "P10:R18")



#    clear_border(ws, 'S19:S19')
 #   clear_border(ws, 'A19:R19')
    #ws['B7'].borders.top.border_style = None
    #clear_border(ws, 'A'+str(c-1)+':S'+str(c-1))
    #clear_border(ws, 'R'+str(c-1)+':R'+str(c-1))
    #clear_border(ws, 'S'+str(c-1)+':S'+str(c-1))
    #set_border(ws, 'S'+str(c-2)+':S'+str(c-2))
    #set_border(ws, 'B'+str(c-2)+':N'+str(c-2))
    #set_border(ws, 'P'+str(c-2)+':R'+str(c-2))

    sf_name = 'ttt'+'.xlsx'
    try:
        os.remove('blast_report/'+sf_name)
    except:
        pass
    wb.save('blast_report/'+sf_name)

    return

test()
exit()



print time.gmtime()
print time.localtime()
print timezone.now()

a = 100
b = 0

s = a if b else 0
print s

exit()


def today():
    current = datetime.datetime.now()
#    current = timezone.now()
    return current.strftime('%Y-%m-%d')


print today()


def now():
    current = datetime.datetime.now()
#    current = timezone.now()
    return current.strftime('%Y-%m-%d %H:%M:%S')

print now()

exit()

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]



# getDownloadFingerServerUri()
#     getServer1Address() + "/service/v1/scanner/company_burst_person/fingers";
# http get
# http://182.92.190.30:9210/service/v1/scanner/company_burst_person/fingers?data=%7B%22equip_id%22%3A%22861189014119126%22%2C%22company_id%22%3A%22159e42ae569411e5b713dea90594582e%22%7D
# data={"equip_id":"861189014119126","company_id":"159e42ae569411e5b713dea90594582e"}
# referring to getDownloadFingerServerUri.json
# {
#   "user_fingers": [
#     {
#       "user_name": "吴五",
#       "user_id": "10d87cc0619611e58855dea90594582e",
#       "user_type": 3,
#       "fingers": [
#         {
#           "id": "7f5b71382bb341878f66dd3469576a66",
#           "value": "ywkHZqGq+pTnAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA==\n"
#         }
#       ]
#     },
#     {
#       "user_name": "盛爆2",
#       "user_id": "1dacdf00659111e594a3dea90594582e",
#       "user_type": 3,
#       "fingers": [
#         {
#           "id": "26e4d1cc55ee4dc5a02c0c61739cee71",
#           "value": "ywkPYqGqupRFR75wTW1CbKqbDHi7thR0l2wccLfmpJyr+dnxeGp7M33EAl8Cb15sOJrfY+msLHuj\ncAxRYa6kM3AZqGZK6smC90OfjOojn/T8hn6q7oizZjEw+mYHTHmdaZ+1Nq9atlYAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA==\n"
#         },
#         {
#           "id": "33980497f37c4dbdb36a950b94583814",
#           "value": "ywkOXKGqmhTFxv5gze0CdKsbDHC7N5R4lW+cfLTg5ICoxbkJKItrV/REulUibvZuBJQUdGa4PCzZ\nEl8eDW60e0A4/UgP6Fn2g064Ahl+Lw7yvFySOtB9sG24vKi6XFoLWSm1NgAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA==\n"
#         },
#         {
#           "id": "7bcd786293554724bd954507699fc297",
#           "value": "ywkPZKGqhZQ7RYGQMW56fGobzHQ6N5R4lW+cfLXgJICvxVgJ75uCTIhF33wwbZdD5ZxGRDS4rVgq\nUB9M6L6HxVIZngxDKe8I/sNZRHZmmcvThlzRyYXzTqURklxX/fzbFui1Nk6ytlYAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\nAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA==\n"
#         }
#       ]
#     }
#   ],
#   "version": 1,
#   "success": true
# }

def mfingers_company(company_id):
    dict = {}
    dict['version'] = 1
    dict['success'] = True
    dict['user_fingers'] = []



    cursor = connection.cursor()
    sql = 'SELECT * FROM view_finger WHERE company_id=\'' +company_id+ '\''
    cursor.execute(sql)
    results = dictfetchall(cursor)

    user_fingers = {}
    for result in results:
        user_id = result['user_id']
        finger = {"id":result['id'], "value":result['finger_data']}

        if user_id in user_fingers.keys():
            user_fingers[user_id]["fingers"].append(finger)
        else:
            user_fingers[user_id] = {
                "user_name":result['user_name'],
                "user_id":user_id,
                "user_type":result['user_type'],
                "fingers":[finger]
            }

    for each in user_fingers:
        dict['user_fingers'].append(user_fingers[each])

    return json.dumps(dict)


print mfingers_company('e42bd46740034ede8ac7ab27416368a4')

