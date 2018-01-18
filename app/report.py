# coding=utf-8

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Alignment
import django
django.setup()

from django.db import connection
from django.http import HttpResponse
from models import *
import json
import os
import snippet


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

def string_reverse1(string):
    return string[::-1]

def double(req):
    double_ids = req.GET.get('id', '')

    print double_ids
    print '1'
    f_name = 'template/blast_record_210000.xlsx'
    wb = load_workbook(filename=f_name)
    ws = wb.active

    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM')
    date_style.alignment = Alignment(horizontal='general', vertical='center')
    wrap_alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True)

    cursor = connection.cursor()
    record_start = 4

    double_ID = string_reverse1(double_ids)

    IDS = []
    t = 0
    ids = 0
    print double_ID
    for a in double_ID:
        if str(a) != ',':
            temp = int(a)
            ids = ids + temp * pow(10,t)
            t = t+1
        else:
            t = 0
            IDS.append(ids)
            ids = 0
    IDS.append(ids)

    IDS_Len = len(IDS)

    set_border(ws, "A1:A" + str(IDS_Len + 3))
    set_border(ws, "B1:B" + str(IDS_Len + 3))
    set_border(ws, "C1:C" + str(IDS_Len + 3))
    set_border(ws, "D1:D" + str(IDS_Len + 3))
    set_border(ws, "E1:E" + str(IDS_Len + 3))
    set_border(ws, "F1:F" + str(IDS_Len + 3))
    set_border(ws, "G1:G" + str(IDS_Len + 3))
    set_border(ws, "H1:H" + str(IDS_Len + 3))
    set_border(ws, "I1:I" + str(IDS_Len + 3))


    for stu_id in IDS:
        sql_dgt = '''
            SELECT a.id, a.user_id, a.name, a.sex_id, b.sex, h.major,a.stu_class,institution_name, g.name AS Bname, f.subject,
             f.research_papers,f.engineering_design,f.project_report,f.summary_report,f.other,f.other_introduction,
             f.combine_actual,f.company_name,f.implementation_state,f.english_title,f.subject_property
            FROM app_select_sysuser a
                LEFT JOIN app_sex b ON a.sex_id = b.id
                LEFT JOIN app_major h ON a.major_id = h.id
                LEFT JOIN app_application_state d ON a.user_id = d.selected_stu_no
                LEFT JOIN app_stu_introduction e ON a.user_id = e.stu_no
                LEFT JOIN app_topic f ON d.topic = f.id
                LEFT JOIN app_select_sysuser AS g ON f.tea_no = g.user_id
                LEFT JOIN app_institution c ON g.institution_id = c.id
            WHERE a.id=
            '''

        sql_dgt += '\'' + str(stu_id) + '\''
        cursor.execute(sql_dgt)


        sql_dgt_out = dictfetchall(cursor)
        dgt_count = len(sql_dgt_out)

        if dgt_count == 1 and sql_dgt_out[0]['id'] == None:
            print 'No used matertial record for area ' + stu_id
        else:
            n = str(record_start)
            ws['A' + n] = record_start-3
            ws['B' + n] = sql_dgt_out[0]['name']
            ws['C' + n] = sql_dgt_out[0]['user_id']
            ws['D' + n] = sql_dgt_out[0]['sex']
            ws['E' + n] = sql_dgt_out[0]['major']
            ws['F' + n] = sql_dgt_out[0]['stu_class']
            ws['G' + n] = sql_dgt_out[0]['Bname']
            ws['H' + n] = sql_dgt_out[0]['institution_name']
            ws['I' + n] = sql_dgt_out[0]['subject']

            ws['S' + n] = sql_dgt_out[0]['english_title']
            ws['J' + n] = sql_dgt_out[0]['research_papers']
            ws['K' + n] = sql_dgt_out[0]['engineering_design']
            ws['L' + n] = sql_dgt_out[0]['project_report']
            ws['M' + n] = sql_dgt_out[0]['summary_report']
            ws['N' + n] = sql_dgt_out[0]['other']
            ws['O' + n] = sql_dgt_out[0]['other_introduction']
            ws['P' + n] = sql_dgt_out[0]['combine_actual']
            ws['Q' + n] = sql_dgt_out[0]['company_name']
            ws['R' + n] = sql_dgt_out[0]['implementation_state']

            record_start += 1

    try:
        os.remove('double_report.xlsx')
    except:
        pass
    wb.save('double_report.xlsx')

    return json.dumps({'report_name': 'double_report.xlsx'})

def title(req):
    stu_major = req.GET.get('stuMajor', '')
    print 'stu_major'+stu_major
    f_name = 'download/summary_title.xlsx'
    wb = load_workbook(filename=f_name)
    ws = wb.active
    cursor = connection.cursor()
    record_start = 4

    sql = '''SELECT a.id,a.subject,f.institution_name,b.name,a.introduction FROM app_topic a
                     LEFT JOIN app_select_sysuser b ON a.tea_no=b.user_id
                     LEFT JOIN app_institution f ON f.id = b.institution_id
                     LEFT JOIN app_date_setting c ON a.year=c.year
                     WHERE a.id not in (SELECT app_application_state.topic FROM app_application_state)
                     AND c.activation=1 
                     AND a.major_id=''' + stu_major
    cursor.execute(sql)
    sql_out = dictfetchall(cursor)
    dgt_count = len(sql_out)
    print dgt_count
    print sql_out

    for so in sql_out:
        n = str(record_start)
        ws['A' + n] = record_start - 3
        ws['B'+n] = so['name']
        ws['C'+n] = so['institution_name']
        ws['D'+n] = so['subject']
        ws['E'+n] = so['introduction']
        record_start += 1
    print 'title 结束'
    try:
        os.remove('summary_title.xlsx')
    except:
        pass
    wb.save('summary_title.xlsx')
    return json.dumps({'report_name': 'summary_title.xlsx'})


def blastarea(req):
    area_id = req.GET.get('id', '')
    area = Area.objects.get(id=area_id)
    print 'generate blast report for area ' + str(area_id)

    f_name = 'template/blast_record_210000.xlsx'
    wb = load_workbook(filename=f_name)
    ws = wb.active

    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM')
    date_style.alignment = Alignment(horizontal='general', vertical='center')
    wrap_alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True)

    set_border(ws, "A3:S8")
    set_border(ws, "R3:R4")
    set_border(ws, "B11:R11")
    set_border(ws, "B13:R13")
    set_border(ws, "B15:R15")
    set_border(ws, "B17:R17")
    set_border(ws, "B10:N18")
    set_border(ws, "P10:R18")

    ws['C2'] = area.name
    wb.guess_types = True
    ws['Q2'] = area.describe
    ws['B5'] = area.plan_explosive_num
    ws['C5'] = area.plan_dgt_deto_num
    ws['D5'] = area.plan_deto_num

    cursor = connection.cursor()
    record_start = 11

    sql_dgt = '''
    SELECT area_id, blast_time, normal_count, abnormal_count
    FROM view_digitalstat
    WHERE area_id=
    '''
    sql_dgt += '\'' + area_id + '\''
    cursor.execute(sql_dgt)

    sql_dgt_out = dictfetchall(cursor)
    dgt_count = len(sql_dgt_out)
    if dgt_count == 1 and sql_dgt_out[0]['area_id'] == None:
        print 'No used matertial record for area ' + area_id
    else:
        if dgt_count > 3:
            record_start += dgt_count - 3
            ws.insert_rows(8, dgt_count - 3, above=True, copy_style=True)
        if dgt_count > 0:
            br_index = 0
            for br in sql_dgt_out:
                N = str(5 + br_index)
                ws['A'+N] = br['blast_time']
                ws['A'+N].style = date_style
                set_border(ws, 'A'+N+':A'+N)
                ws['H'+N] = br['normal_count']
                ws['M'+N] = br['abnormal_count']
                br_index += 1

    blastrecords = BlastRecord.objects.filter(area_id=area_id)
    br_count = len(blastrecords)
    if br_count > (3-dgt_count):
        record_start += br_count-(3-dgt_count)
        ws.insert_rows(5+dgt_count, br_count-(3-dgt_count), above=False, copy_style=True)
    if br_count > 0:
        br_index = 0
        for br in blastrecords:
            N = str(5 + dgt_count + br_index)
            ws['A'+N] = br.blasting_time
            ws['A'+N].style = date_style
            set_border(ws, 'A'+N+':A'+N)
            try:
                user1 = SysUser.objects.get(id=br.user1_id)
                ws['Q'+N] = user1.user_name
            except:
                pass
            try:
                user2 = SysUser.objects.get(id=br.user2_id)
                ws['R'+N] = user2.user_name
            except:
                pass
            br_index += 1

    sql = '''
        SELECT area_id,
	        SUM(CASE WHEN resource_sub_type_id IN (2,13) THEN resource_count ELSE 0 END) AS dgt_deto_count,
	        SUM(CASE WHEN resource_sub_type_id IN (3,7,10) THEN resource_count ELSE 0 END) AS ele_deto_count,
	        SUM(CASE WHEN resource_sub_type_id IN (4,5,6,8,9,11,12) THEN resource_count ELSE 0 END) AS tube_deto_count,
	        SUM(CASE WHEN resource_sub_type_id IN (14,15,16,17) THEN resource_count ELSE 0 END) AS explosive_count,
	        SUM(CASE WHEN resource_sub_type_id IN (14) THEN resource_count ELSE 0 END) AS explosive_count_02W,
	        SUM(CASE WHEN resource_sub_type_id IN (15) THEN resource_count ELSE 0 END) AS explosive_count_034,
	        SUM(CASE WHEN resource_sub_type_id IN (16) THEN resource_count ELSE 0 END) AS explosive_count_03W,
	        SUM(CASE WHEN resource_sub_type_id IN (17) THEN resource_count ELSE 0 END) AS explosive_count_03X,
	        SUM(CASE WHEN resource_sub_type_id IN (19) THEN resource_count ELSE 0 END) AS nonel_deto_count,
	        SUM(CASE WHEN resource_sub_type_id IN (18) THEN resource_count ELSE 0 END) AS fuse_count
	        FROM view_materialrecord
	'''

    sql_used = sql + ' WHERE record_type_id=1 AND area_id=\'' + area_id + '\''
    cursor.execute(sql_used)

    sql_used_out = dictfetchall(cursor)[0]
    if sql_used_out['area_id'] == None:
        print 'No used matertial record for area ' + area_id
    else:
        suo_id = 0
        if sql_used_out['explosive_count_02W'] != 0:
            ws['G' + str(5+dgt_count+suo_id)] = '乳化炸药\n' + str(sql_used_out['explosive_count_02W']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_034'] != 0:
            ws['G' + str(5+dgt_count+suo_id)] = '粉状\n' + str(sql_used_out['explosive_count_034']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_03W'] != 0:
            ws['G' + str(5+dgt_count+suo_id)] = '改性铵油\n' + str(sql_used_out['explosive_count_03W']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_03X'] != 0:
            ws['G' + str(5+dgt_count+suo_id)] = '胶质\n' + str(sql_used_out['explosive_count_03X']).rstrip('0').rstrip('.')
            suo_id += 1

        ws['H5'] = sql_used_out['dgt_deto_count']
        ws['I5'] = sql_used_out['tube_deto_count']
        ws['J5'] = sql_used_out['nonel_deto_count']
        ws['K5'] = sql_used_out['fuse_count']

    sql_used = sql + ' WHERE record_type_id=4 AND area_id=\'' + area_id + '\''
    cursor.execute(sql_used)

    sql_used_out = dictfetchall(cursor)[0]
    if sql_used_out['area_id'] == None:
        print 'No recalled matertial record for area ' + area_id
    else:
        suo_id = 0
        if sql_used_out['explosive_count_02W'] != 0:
            ws['L' + str(5+dgt_count+suo_id)] = '乳化炸药\n' + str(sql_used_out['explosive_count_02W']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_034'] != 0:
            ws['L' + str(5+dgt_count+suo_id)] = '粉状\n' + str(sql_used_out['explosive_count_034']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_03W'] != 0:
            ws['L' + str(5+dgt_count+suo_id)] = '改性铵油\n' + str(sql_used_out['explosive_count_03W']).rstrip('0').rstrip('.')
            suo_id += 1
        if sql_used_out['explosive_count_03X'] != 0:
            ws['L' + str(5+dgt_count+suo_id)] = '胶质\n' + str(sql_used_out['explosive_count_03X']).rstrip('0').rstrip('.')
            suo_id += 1

        ws['M5'] = sql_used_out['dgt_deto_count']
        ws['N5'] = sql_used_out['tube_deto_count']
        ws['O5'] = sql_used_out['nonel_deto_count']
        ws['P5'] = sql_used_out['fuse_count']

    sql_used = '''
    SELECT area_id, user_name, simple_code, start_index, end_index, resource_count
	FROM view_materialrecord
	WHERE record_type_id=1 AND resource_type_id=0 AND area_id=
    '''
    sql_used += '\'' + area_id + '\''
    cursor.execute(sql_used)

    sql_used_out = dictfetchall(cursor)

    sql_used = '''
        SELECT area_id, user_name, simple_code, start_index, end_index, resource_count
        FROM view_materialrecord
        WHERE record_type_id=4 AND resource_type_id=0 AND area_id=
    '''
    sql_used += '\'' + area_id + '\''
    cursor.execute(sql_used)

    sql_recall_out = dictfetchall(cursor)

    begin = record_start
    out_count = 0

    for each in sql_used_out:
        if each['area_id'] == None:
            print 'No used matertial record for area ' + area_id
        else:
            find_out = False
            find_id = 0
            for a in range(out_count):
                ws['A' + str(begin + a)].value == each['user_name']
                find_out = True
                find_id = a
                break

            if find_out == False:
                if out_count > 6:
                    ws.insert_rows(begin+out_count, 1, above=True, copy_style=True)
                ws['A'+str(record_start)] = each['user_name']
                ws['B'+str(record_start)] = each['simple_code'] + '(' +str(each['start_index'])+ '-' +str(each['end_index'])+ ')'
                ws['O'+str(record_start)] = each['resource_count']
                record_start += 1
                out_count += 1
            else:
                N = str(begin+find_id)
                ws['B'+N] = ws['B'+N].value + ', ' + \
                    each['simple_code'] + '(' +str(each['start_index'])+ '-' +str(each['end_index'])+ ')'
                ws['O'+N] = ws['O'+N].value + each['resource_count']

    for each in sql_recall_out:
        if each['area_id'] == None:
            print 'No recalled matertial record for area ' + area_id
        else:
            find_out = False
            find_id = 0
            for a in range(out_count):
                ws['A' + str(begin + a)].value == each['user_name']
                find_out = True
                find_id = a
                break

            if find_out == False:
                if out_count > 6:
                    ws.insert_rows(begin+out_count, 1, above=True, copy_style=True)
                ws['A'+str(record_start)] = each['user_name']
                ws['P'+str(record_start)] = each['simple_code'] + '(' +str(each['start_index'])+ '-' +str(each['end_index'])+ ')'
                ws['S'+str(record_start)] = each['resource_count']
                record_start += 1
                out_count += 1
            else:
                N = str(begin+find_id)
                ws['P'+N] = ws['P'+N].value + ', ' + \
                    each['simple_code'] + '(' +str(each['start_index'])+ '-' +str(each['end_index'])+ ')'
                ws['S'+N] = ws['S'+N].value + each['resource_count']

    sql_dgt = '''
    SELECT area_id, code, status_id
    FROM view_digitaldetonator
    WHERE area_id=
    '''
    sql_dgt += '\'' + area_id + '\''
    cursor.execute(sql_dgt)
    sql_dgt_out = dictfetchall(cursor)

    first_normal = 1
    first_abnormal = 1
    for each in sql_dgt_out:
        if each['area_id'] == None:
            print 'No recalled matertial record for area ' + area_id
        else:
            if out_count > 6:
                ws.insert_rows(begin+out_count, 1, above=True, copy_style=True)

            if each['status_id'] == 0:
                if first_normal == 1:
                    ws['B'+str(record_start)] = str(each['code'])
                    ws['O'+str(record_start)] = 1
                    first_normal = 0
                else:
                    ws['B' + str(record_start)] = ws['B' + str(record_start)].value + ', '+ str(each['code'])
                    ws['O' + str(record_start)] = int(ws['O' + str(record_start)].value) + 1
            else:
                if first_abnormal == 1:
                    ws['P'+str(record_start)] = str(each['code'])
                    ws['S'+str(record_start)] = 1
                    first_abnormal = 0
                else:
                    ws['P' + str(record_start)] = ws['P' + str(record_start)].value + ', '+ str(each['code'])
                    ws['S' + str(record_start)] = int(ws['S' + str(record_start)].value) + 1

    ws['P' + str(record_start)].alignment = wrap_alignment

    sf_name = area_id+'.xlsx'
    try:
        os.remove('blast_report/'+sf_name)
    except:
        pass
    wb.save('blast_report/'+sf_name)

    return json.dumps({'report_name': sf_name})


def report(req):
    report_type = req.GET.get('type', '')

    if report_type == 'area':
        ret = blastarea(req)
        return HttpResponse(ret)

    return HttpResponse('-1')


def double_report(req):
    report_type = req.GET.get('type', '')

    if report_type == 'double':
        ret = double(req)

        return HttpResponse(ret)

    return HttpResponse('-1')

def summary_title(req):
    report_type = req.GET.get('type', '')
    print 'here'
    print report_type
    if report_type == 'title':
        ret = title(req)
        return HttpResponse(ret)

    return HttpResponse('-1')